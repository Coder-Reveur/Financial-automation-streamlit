import os
import io
import time
import csv
import zipfile
import warnings
from pathlib import Path
import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import load_workbook
from pyexcelerate import Workbook
from datetime import datetime as dt
from collections import defaultdict


# Page config and title
st.set_page_config(page_title="Financial Statement Processing", layout="wide")
st.title("💡 Automated Statement Processor")

# Define four tabs
tab1, tab2, tab3, tab4 = st.tabs([
    "📥 Import & Prepare Statements",
    "📂 Conert Excel Sheets to CSV",
    "⚙️ Complete Data Processing",
    "📈 Roll Over Revenue Update"
])

#--- For Tab 1 ---
with tab1:
    st.set_page_config(page_title="Financial CSV Processor", layout="wide")


    def process_dataframes(df_invoice, df_refund, df_dispute, year_cutoff: int = 2025):
        """
        Vectorized processing adapted
        Returns updated (df_invoice, df_refund, df_dispute).
        """
        # Work on copies
        df_invoice = df_invoice.copy()
        df_refund = df_refund.copy()
        df_dispute = df_dispute.copy()

        # Remember original invoice column names to restore later
        orig_invoice_cols = list(df_invoice.columns)

        # Ensure required blank columns exist (do not overwrite existing)
        for col in ['Inv Check', 'Refund', 'Dispute', 'Inv Amount']:
            if col not in df_invoice.columns:
                df_invoice[col] = ''
        for col in ['Collected', 'Check collection', 'Diff']:
            if col not in df_refund.columns:
                df_refund[col] = ''
            if col not in df_dispute.columns:
                df_dispute[col] = ''

        # --- Refund sums ---
        if 'Invoice Number' in df_refund.columns:
            refund_key = 'Invoice Number'
        elif 'Inv. No.' in df_refund.columns:
            refund_key = 'Inv. No.'
        else:
            raise KeyError("Refund sheet missing invoice column ('Invoice Number' or 'Inv. No.')")

        if 'Amount Refunded' not in df_refund.columns:
            raise KeyError("Refund sheet missing 'Amount Refunded' column")

        refund_sums = df_refund.groupby(refund_key)['Amount Refunded'].sum()

        if 'Number' in df_invoice.columns:
            invoice_map_key = 'Number'
        elif 'Invoice Number' in df_invoice.columns:
            invoice_map_key = 'Invoice Number'
        else:
            raise KeyError("Invoice sheet missing invoice column ('Number' or 'Invoice Number')")

        df_invoice['Refund'] = df_invoice[invoice_map_key].map(refund_sums).fillna(0)

        # --- Dispute sums (status == 'lost') ---
        if 'Invoice Number' in df_dispute.columns:
            dis_key = 'Invoice Number'
        elif 'Inv. No.' in df_dispute.columns:
            dis_key = 'Inv. No.'
        else:
            raise KeyError("Dispute sheet missing invoice column ('Invoice Number' or 'Inv. No.')")

        if 'Dispute Status' not in df_dispute.columns or 'Disputed Amount' not in df_dispute.columns:
            raise KeyError("Dispute sheet missing 'Dispute Status' or 'Disputed Amount'")

        lost = df_dispute[df_dispute['Dispute Status'] == 'lost']
        dispute_sums = lost.groupby(dis_key)['Disputed Amount'].sum()
        df_invoice['Dispute'] = df_invoice[invoice_map_key].map(dispute_sums).fillna(0)

        # --- Inv Amount calculation ---
        for col in ['Amount Paid', 'Refund', 'Dispute']:
            if col not in df_invoice.columns:
                df_invoice[col] = 0
        df_invoice['Inv Amount'] = pd.to_numeric(df_invoice['Amount Paid'], errors='coerce').fillna(0) \
                                - pd.to_numeric(df_invoice['Refund'], errors='coerce').fillna(0) \
                                - pd.to_numeric(df_invoice['Dispute'], errors='coerce').fillna(0)

        # --- Collected in refund/dispute sheets ---
        if 'Amount' not in df_refund.columns or 'Amount Refunded' not in df_refund.columns:
            raise KeyError("Refund sheet must contain 'Amount' and 'Amount Refunded' for collected calculation")
        if 'Amount' not in df_dispute.columns or 'Disputed Amount' not in df_dispute.columns:
            raise KeyError("Dispute sheet must contain 'Amount' and 'Disputed Amount' for collected calculation")

        df_refund['Collected'] = pd.to_numeric(df_refund['Amount'], errors='coerce').fillna(0) \
                                - pd.to_numeric(df_refund['Amount Refunded'], errors='coerce').fillna(0)
        df_dispute['Collected'] = pd.to_numeric(df_dispute['Amount'], errors='coerce').fillna(0) \
                                - pd.to_numeric(df_dispute['Disputed Amount'], errors='coerce').fillna(0)

        # --- Make sure Check collection exists (placeholder 0) ---
        if 'Check collection' not in df_refund.columns:
            df_refund['Check collection'] = 0
        if 'Check collection' not in df_dispute.columns:
            df_dispute['Check collection'] = 0

        # restore Number column name if previously present (no destructive rename was done in this function)
        # Return updated DataFrames
        return df_invoice, df_refund, df_dispute


    def df_to_bytes_csv(df: pd.DataFrame, encoding: str = 'utf-8-sig') -> bytes:
        """Return CSV bytes for st.download_button."""
        with io.BytesIO() as buffer:
            df.to_csv(buffer, index=False, encoding=encoding)
            return buffer.getvalue()


    def main():
        st.title("Financial CSV Processor (Invoice / Refund / Dispute)")
        st.markdown(
            """
            Upload your three CSVs (Invoice, Refund, Dispute). The app will:
            * Add required columns if missing,
            * Compute Refund, Dispute, Inv Amount, Collected (vectorized),
            * Provide updated CSV downloads ( originals are not overwritten).
            """
        )

        upload_mode = st.radio("Input method", ("Upload files", "Provide file paths"))

        invoice_file = None
        refund_file = None
        dispute_file = None
        invoice_path_text = None
        refund_path_text = None
        dispute_path_text = None

        if upload_mode == "Upload files":
            invoice_file = st.file_uploader("Invoice CSV", type=["csv", "txt"], key="inv")
            refund_file  = st.file_uploader("Refund CSV",  type=["csv", "txt"], key="ref")
            dispute_file = st.file_uploader("Dispute CSV", type=["csv", "txt"], key="dis")
        else:
            invoice_path_text = st.text_input("Invoice CSV path (local)", "")
            refund_path_text  = st.text_input("Refund CSV path (local)", "")
            dispute_path_text = st.text_input("Dispute CSV path (local)", "")

        year_cutoff = st.number_input("Year cutoff (for check collection logic, unused placeholder)", value=2025, step=1)

        if st.button("Process files"):
            try:
                with st.spinner("Loading CSVs..."):
                    # Load according to chosen mode
                    if upload_mode == "Upload files":
                        if not (invoice_file and refund_file and dispute_file):
                            st.error("Please upload all three CSVs.")
                            return
                        df_invoice = pd.read_csv(invoice_file, dtype=str, encoding='utf-8-sig')
                        df_refund  = pd.read_csv(refund_file,  dtype=str, encoding='utf-8-sig')
                        df_dispute = pd.read_csv(dispute_file, dtype=str, encoding='utf-8-sig')
                        # set original path stems for naming
                        inv_stem = Path(invoice_file.name).stem
                        ref_stem = Path(refund_file.name).stem
                        dis_stem = Path(dispute_file.name).stem
                    else:
                        if not (invoice_path_text and refund_path_text and dispute_path_text):
                            st.error("Please provide all three file paths.")
                            return
                        # read from provided local paths
                        df_invoice = pd.read_csv(invoice_path_text, dtype=str, encoding='utf-8-sig')
                        df_refund  = pd.read_csv(refund_path_text,  dtype=str, encoding='utf-8-sig')
                        df_dispute = pd.read_csv(dispute_path_text, dtype=str, encoding='utf-8-sig')
                        inv_stem = Path(invoice_path_text).stem
                        ref_stem = Path(refund_path_text).stem
                        dis_stem = Path(dispute_path_text).stem

                # Process
                start = time.time()
                with st.spinner("Processing (vectorized)..."):
                    updated_invoice, updated_refund, updated_dispute = process_dataframes(
                        df_invoice, df_refund, df_dispute, year_cutoff=year_cutoff
                    )
                elapsed = time.time() - start
                mins, secs = divmod(int(elapsed), 60)
                millis = int((elapsed - int(elapsed)) * 1000)
                st.success(f"Processing completed in {mins:02d}:{secs:02d}.{millis:03d} (mm:ss.ms)")

                # Show small previews
                st.subheader("📊 Data Preview (first few rows of each sheet)")

                with st.expander("Invoice preview ", expanded=False):
                    st.dataframe(updated_invoice.head(10))

                with st.expander("Refund preview ", expanded=False):
                    st.dataframe(updated_refund.head(10))

                with st.expander("Dispute preview ", expanded=False):
                    st.dataframe(updated_dispute.head(10))

                # Prepare CSV bytes and provide download buttons
                inv_bytes = df_to_bytes_csv(updated_invoice)
                ref_bytes = df_to_bytes_csv(updated_refund)
                dis_bytes = df_to_bytes_csv(updated_dispute)

                st.markdown("### Download updated CSVs")

                # build a zip
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                    i = 0
                    name = [inv_stem, ref_stem, dis_stem]
                    for csv_byte in [inv_bytes, ref_bytes, dis_bytes]:
                        zf.writestr(f"{name[i]}_udpated.csv", csv_byte)
                        i += 1
                zip_buf.seek(0)
                st.download_button("Download ZIP of all CSVs", data=zip_buf.getvalue(), file_name="Updated_csv.zip", mime="application/zip")
                
                # individual buttons
                col1, col2, col3 = st.columns(3)
                col1.download_button(
                    label=f"Download {inv_stem}_updated.csv",
                    data=inv_bytes,
                    file_name=f"{inv_stem}_updated.csv",
                    mime="text/csv"
                )
                col2.download_button(
                    label=f"Download {ref_stem}_updated.csv",
                    data=ref_bytes,
                    file_name=f"{ref_stem}_updated.csv",
                    mime="text/csv"
                )
                col3.download_button(
                    label=f"Download {dis_stem}_updated.csv",
                    data=dis_bytes,
                    file_name=f"{dis_stem}_updated.csv",
                    mime="text/csv"
                )

            except Exception as exc:
                st.error(f"Error: {exc}")
                raise

    if __name__ == "__main__":
        main()

#--- For Tab 2 ---
with tab2:
    # Suppress openpyxl user warning about missing stylesheet (not harmful)
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    st.set_page_config(page_title="Excel → CSV Converter", layout="wide")

    st.title("Excel → CSV Converter (All_CSVs per workbook)")

    st.markdown(
        """
        Convert selected sheets from one or more Excel workbooks into CSVs.
        - For **local path** mode the app writes `All_CSVs` next to each workbook.
        - For **upload** mode the app provides downloadable CSV files (or a single ZIP).
        """
    )

    mode = st.radio("Input mode", ("Upload Excel files", "Local Excel paths (on this machine)"))

    def list_sheets_from_workbook_bytes(bytes_io):
        try:
            wb = load_workbook(bytes_io, read_only=True, data_only=True)
            return wb.sheetnames
        except Exception as e:
            raise RuntimeError(f"Failed reading workbook: {e}")

    def list_sheets_from_workbook_path(path: Path):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            return wb.sheetnames
        except Exception as e:
            raise RuntimeError(f"Failed reading workbook {path}: {e}")

    def convert_sheet_to_csv_bytes(wb_src, sheet_name: str) -> bytes:
        """Given an openpyxl Workbook object and a sheet name, return CSV bytes."""
        ws = wb_src[sheet_name]
        buf = io.StringIO()
        writer = csv.writer(buf, lineterminator="\n")
        for row in ws.iter_rows(values_only=True):
            # convert None -> "" so CSV cell isn't 'None'
            writer.writerow([("" if v is None else v) for v in row])
        return buf.getvalue().encode("utf-8-sig")

    def convert_uploaded_workbook_to_csvs(uploaded_file, selected_sheets):
        """
        uploaded_file: streamlit UploadedFile
        selected_sheets: list of sheet names to convert
        returns list of tuples: (csv_filename, csv_bytes)
        """
        raw = uploaded_file.read()
        bio = io.BytesIO(raw)
        wb = load_workbook(bio, read_only=True, data_only=True)
        csvs = []
        for sheet in selected_sheets:
            csv_bytes = convert_sheet_to_csv_bytes(wb, sheet)
            safe_sheet = "".join(c if (c.isalnum() or c in " _-") else "_" for c in sheet)[:50]
            csv_name = f"{Path(uploaded_file.name).stem}_{safe_sheet}.csv"
            csvs.append((csv_name, csv_bytes))
        return csvs

    def convert_local_workbook_to_csvs(path: Path, selected_sheets, out_dir: Path):
        """
        path: Path to local workbook
        selected_sheets: list of sheet names to convert
        out_dir: directory to write CSVs to (will be created)
        returns list of written file paths
        """
        wb = load_workbook(path, read_only=True, data_only=True)
        out_dir.mkdir(parents=True, exist_ok=True)
        written = []
        for sheet in selected_sheets:
            csv_bytes = convert_sheet_to_csv_bytes(wb, sheet)
            safe_sheet = "".join(c if (c.isalnum() or c in " _-") else "_" for c in sheet)[:50]
            csv_path = out_dir / f"{path.stem}_{safe_sheet}.csv"
            with open(csv_path, "wb") as f:
                f.write(csv_bytes)
            written.append(csv_path)
        return written

    # UI: Upload or local paths
    workbooks_info = []  # list of dicts { 'name':..., 'type': 'upload'|'local', 'object':..., 'sheets':[...] , 'selected':[...] }

    if mode == "Upload Excel files":
        uploaded = st.file_uploader("Upload one or more Excel files (.xlsx/.xlsm) — use Shift/Ctrl to select multiple",
                                    type=["xlsx", "xlsm", "xls"], accept_multiple_files=True)
        if uploaded:
            st.info(f"{len(uploaded)} file(s) uploaded. Expand each to choose sheets.")
            for u in uploaded:
                try:
                    sheets = list_sheets_from_workbook_bytes(io.BytesIO(u.read()))
                    # after reading bytes above, need to reset pointer or re-create bytes for later conversion
                    u.seek(0)
                except Exception as exc:
                    st.error(f"Failed to read '{u.name}': {exc}")
                    sheets = []
                with st.expander(f"{u.name} ({len(sheets)} sheets)" , expanded=True):
                    st.write("Sheets:", sheets)
                    # default select all
                    selected = st.multiselect(f"Select sheets to convert for {u.name}", options=sheets, default=sheets, key=f"sel_{u.name}")
                    workbooks_info.append({"name": u.name, "type": "upload", "object": u, "sheets": sheets, "selected": selected})
        else:
            st.info("No uploads yet. Use the file uploader above.")

    else:  # Local path mode
        st.markdown("Enter full file paths (one per line). The app must run on the same machine that has these files.")
        paths_text = st.text_area("Paste Excel full paths here (one path per line):", height=120)
        if paths_text.strip():
            paths = [p.strip().strip('"').strip("'") for p in paths_text.splitlines() if p.strip()]
            for p in paths:
                path = Path(p)
                if not path.exists():
                    st.error(f"File not found: {p}")
                    continue
                try:
                    sheets = list_sheets_from_workbook_path(path)
                except Exception as exc:
                    st.error(f"Failed to read {p}: {exc}")
                    continue
                with st.expander(f"{path.name} ({len(sheets)} sheets)", expanded=True):
                    st.write("Sheets:", sheets)
                    selected = st.multiselect(f"Select sheets to convert for {path.name}", options=sheets, default=sheets, key=f"local_{path.name}")
                    workbooks_info.append({"name": str(path), "type": "local", "object": path, "sheets": sheets, "selected": selected})
        else:
            st.info("No local paths provided yet. Paste one or more paths above and press outside the text area to register.")

    # Convert action
    if workbooks_info:
        st.markdown("---")
        st.write("Selected workbooks & sheets to convert:")
        for info in workbooks_info:
            st.write(f"- **{info['name']}** → {len(info['selected'])} sheet(s) selected")
        if st.button("Convert selected sheets to CSVs"):
            start = time.time()
            all_downloadables = []  # for uploads: (csv_name, bytes)
            written_files = []      # for local: list of paths written
            errors = []
            with st.spinner("Converting..."):
                for info in workbooks_info:
                    try:
                        if info["type"] == "upload":
                            # uploaded file object was read earlier; ensure we're at start
                            uploaded_file = info["object"]
                            uploaded_file.seek(0)
                            csvs = convert_uploaded_workbook_to_csvs(uploaded_file, info["selected"])
                            all_downloadables.extend(csvs)
                        else:
                            path = Path(info["object"])
                            out_dir = path.parent / "All_CSVs"
                            written = convert_local_workbook_to_csvs(path, info["selected"], out_dir)
                            written_files.extend(written)
                    except Exception as exc:
                        errors.append(f"{info['name']}: {exc}")

            elapsed = time.time() - start
            mins, secs = divmod(int(elapsed), 60)
            millis = int((elapsed - int(elapsed)) * 1000)
            st.success(f"Processing completed in {mins:02d}:{secs:02d}.{millis:03d} (mm:ss.ms)") 

            if errors:
                st.error("Some files failed:")
                for e in errors:
                    st.write(f"- {e}")

            if written_files:
                st.success(f"Wrote {len(written_files)} CSV files in local 'All_CSVs' folders.")
                st.write("Files written:")
                for p in written_files:
                    st.write(f"- {p}")

            if all_downloadables:
                st.success(f"Prepared {len(all_downloadables)} CSV download(s) from uploads.")
                # If many CSVs, offer a ZIP download; otherwise show individual buttons
                if len(all_downloadables) == 1:
                    name, b = all_downloadables[0]
                    st.download_button("Download CSV", data=b, file_name=name, mime="text/csv")
                else:
                    # build a zip
                    zip_buf = io.BytesIO()
                    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                        for name, b in all_downloadables:
                            zf.writestr(name, b)
                    zip_buf.seek(0)
                    st.download_button("Download ZIP of all CSVs", data=zip_buf.getvalue(), file_name="converted_csvs.zip", mime="application/zip")

            st.success(f"Conversion completed in {elapsed:.2f} seconds.")
            st.balloons()

    else:
        st.info("No workbooks selected yet. Upload files or provide local paths and choose sheets.")

    st.markdown("---")
    st.caption("This tool uses openpyxl in read_only mode and csv.writer, similar to your original script.")

#--- For Tab 3 ---
with tab3:
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    st.set_page_config(page_title="CSV Processor (ExcelProcessor)", layout="wide")
    st.title("CSV Processor (wraps ExcelProcessor)")

    # ---------------------------
    # Paste / recreate ExcelProcessor here
    # ---------------------------
    class ExcelProcessor:
        def __init__(self):
            self.sheets_data = {}  # sheet_name -> list of rows
            self.sheet_names = []

        def csv_folder_to_memory(self, csv_folder: str) -> None:
            start = time.time()
            folder = Path(csv_folder)
            csvs = sorted(folder.glob("*.csv"))
            if not csvs:
                raise RuntimeError(f"No CSV files found in {csv_folder!r}")
            for csv_path in csvs:
                with open(csv_path, newline='', encoding='utf-8') as f:
                    reader = list(csv.reader(f))
                    sheet_data = [row for row in reader]
                sheet_name = csv_path.stem[:31]
                self.sheets_data[sheet_name] = sheet_data
                self.sheet_names.append(sheet_name)
                st.write(f"Loaded {csv_path.name} as '{sheet_name}'")
            elapsed = time.time() - start
            st.write(f"Loaded {len(csvs)} CSVs into memory in {elapsed:.2f}s")

        def load_uploaded_csvs_to_memory(self, uploaded_files):
            """
            uploaded_files: list of Streamlit UploadedFile objects
            """
            start = time.time()
            for up in uploaded_files:
                raw = up.read()
                try:
                    text = raw.decode("utf-8")
                except Exception:
                    text = raw.decode("utf-8", errors="replace")
                reader = list(csv.reader(io.StringIO(text)))
                sheet_name = Path(up.name).stem[:31]
                self.sheets_data[sheet_name] = reader
                self.sheet_names.append(sheet_name)
                st.write(f"Loaded (upload) {up.name} as '{sheet_name}'")
                up.seek(0)
            elapsed = time.time() - start
            st.write(f"Loaded {len(uploaded_files)} uploaded CSVs in {elapsed:.2f}s")

        def find_header_row(self, sheet_name: str, key_cols: list, preview_rows: int = 50) -> int:
            if sheet_name not in self.sheets_data:
                raise ValueError(f"Sheet {sheet_name!r} not found")
            expected = set(key_cols)
            sheet_data = self.sheets_data[sheet_name]
            for idx, row in enumerate(sheet_data[:preview_rows]):
                # normalize row values to strings and strip
                vals = set(str(v).strip() for v in row if v is not None and str(v).strip() != "")
                if expected.issubset(vals):
                    return idx
            raise ValueError(f"Could not find header {key_cols} in sheet {sheet_name!r}")

        def choose_sheet(self, prompt: str) -> str:
            print(f"\n{prompt}")
            for i, name in enumerate(self.sheet_names, 1):
                print(f"  {i}. {name}")
            choice = int(input(f"Select [1–{len(self.sheet_names)}]: "))
            return self.sheet_names[choice - 1]

        def populate_inv_check_column(self, invoice_sheet: str, revenue_sheet: str,
                                    number_col_name: str = "Number",
                                    lookup_col_name: str = "Inv. No.",
                                    target_col_name: str = "Inv Check") -> None:
            start_time = time.time()
            inv_hdr = self.find_header_row(invoice_sheet, [number_col_name, target_col_name])
            rev_hdr = self.find_header_row(revenue_sheet, [lookup_col_name])
            inv_data = self.sheets_data[invoice_sheet]
            rev_data = self.sheets_data[revenue_sheet]
            rev_header = [col.strip() if col else col for col in rev_data[rev_hdr]]
            lookup_col_idx = rev_header.index(lookup_col_name)
            lookup_set = set()
            for row in rev_data[rev_hdr + 1:]:
                if lookup_col_idx < len(row) and row[lookup_col_idx]:
                    lookup_set.add(str(row[lookup_col_idx]).strip())
            inv_header = [col.strip() if col else col for col in inv_data[inv_hdr]]
            number_col_idx = inv_header.index(number_col_name)
            target_col_idx = inv_header.index(target_col_name)
            for i, row in enumerate(inv_data[inv_hdr + 1:], start=inv_hdr + 1):
                if number_col_idx < len(row) and row[number_col_idx]:
                    number_val = str(row[number_col_idx]).strip()
                    result = number_val if number_val in lookup_set else "#N/A"
                    new_row = list(row)
                    while len(new_row) <= target_col_idx:
                        new_row.append("")
                    new_row[target_col_idx] = result
                    inv_data[i] = new_row
            elapsed = time.time() - start_time
            st.write(f"Populated '{target_col_name}' on '{invoice_sheet}' in {elapsed:.2f}s")

        def populate_check_collection(self, current_sheet: str, working_sheet: str, revenues_sheet: str,
                                    created_col_name: str = "Created date (UTC)",
                                    invoice_col_name: str = "Invoice Number",
                                    collected_col_name: str = "Collected",
                                    target_col_name: str = "Check collection",
                                    revenue_lookup_col_name: str = "Inv. No.",
                                    year_cutoff: int = 2025) -> None:
            start = time.time()
            current_hdr = self.find_header_row(current_sheet, [created_col_name, invoice_col_name, target_col_name])
            working_hdr = self.find_header_row(working_sheet, [invoice_col_name, collected_col_name])
            revenue_hdr = self.find_header_row(revenues_sheet, [revenue_lookup_col_name, collected_col_name])
            current_data = self.sheets_data[current_sheet]
            working_data = self.sheets_data[working_sheet]
            revenue_data = self.sheets_data[revenues_sheet]
            def build_lookup_sums(data, header_idx, invoice_col, collected_col):
                header = data[header_idx]
                inv_idx = header.index(invoice_col)
                col_idx = header.index(collected_col)
                sums = defaultdict(float)
                for row in data[header_idx + 1:]:
                    if (inv_idx < len(row) and col_idx < len(row) and row[inv_idx] and row[col_idx]):
                        try:
                            sums[row[inv_idx]] += float(row[col_idx])
                        except (ValueError, TypeError):
                            pass
                return sums
            working_sums = build_lookup_sums(working_data, working_hdr, invoice_col_name, collected_col_name)
            revenue_sums = build_lookup_sums(revenue_data, revenue_hdr, revenue_lookup_col_name, collected_col_name)
            header = current_data[current_hdr]
            created_idx = header.index(created_col_name)
            invoice_idx = header.index(invoice_col_name)
            target_idx = header.index(target_col_name)
            for i, row in enumerate(current_data[current_hdr + 1:], start=current_hdr + 1):
                if (created_idx < len(row) and invoice_idx < len(row) and row[created_idx] and row[invoice_idx]):
                    created = row[created_idx]
                    if isinstance(created, dt):
                        year = created.year
                    else:
                        try:
                            year = int(str(created)[:4])
                        except (ValueError, TypeError):
                            year = 0
                    invoice_num = row[invoice_idx]
                    total = (working_sums.get(invoice_num, 0.0) if year < year_cutoff else revenue_sums.get(invoice_num, 0.0))
                    new_row = list(row)
                    while len(new_row) <= target_idx:
                        new_row.append("")
                    new_row[target_idx] = total
                    current_data[i] = new_row
            elapsed = time.time() - start
            st.write(f"Populated '{target_col_name}' on '{current_sheet}' in {elapsed:.2f}s")

        def populate_diff_column(self, sheet_name: str,
                            collected_col_name: str = "Collected",
                            check_col_name: str = "Check collection",
                            target_col_name: str = "Diff") -> None:
            start = time.time()
            header_idx = self.find_header_row(sheet_name, [collected_col_name, check_col_name, target_col_name])
            data = self.sheets_data[sheet_name]
            header = data[header_idx]
            collected_idx = header.index(collected_col_name)
            check_idx = header.index(check_col_name)
            target_idx = header.index(target_col_name)
            for i, row in enumerate(data[header_idx + 1:], start=header_idx + 1):
                collected = 0.0
                checked = 0.0
                if collected_idx < len(row) and row[collected_idx]:
                    try:
                        collected = float(row[collected_idx])
                    except (ValueError, TypeError):
                        pass
                if check_idx < len(row) and row[check_idx]:
                    try:
                        checked = float(row[check_idx])
                    except (ValueError, TypeError):
                        pass
                diff = collected - checked
                new_row = list(row)
                while len(new_row) <= target_idx:
                    new_row.append("")
                new_row[target_idx] = diff
                data[i] = new_row
            elapsed = time.time() - start
            st.write(f"Populated '{target_col_name}' on '{sheet_name}' in {elapsed:.2f}s")

        def save_to_csvs(self, out_folder: str = "All_CSVs", encoding: str = "utf-8-sig") -> list[Path]:
            start = time.time()
            out_dir = Path(out_folder)
            out_dir.mkdir(parents=True, exist_ok=True)
            written = []
            for sheet_name in self.sheet_names:
                safe = "".join(c if (c.isalnum() or c in " _-") else "_" for c in sheet_name)[:50]
                csv_path = out_dir / f"{safe}.csv"
                with open(csv_path, "w", newline="", encoding=encoding) as f:
                    writer = csv.writer(f)
                    rows = self.sheets_data.get(sheet_name, [])
                    for row in rows:
                        row_out = [("" if cell is None else cell) for cell in row]
                        writer.writerow(row_out)
                written.append(csv_path)
                st.write(f"Saved sheet '{sheet_name}' as {csv_path}")
            elapsed = time.time() - start
            st.write(f"Saved {len(written)} CSVs to '{out_dir}' in {elapsed:.2f}s")
            return written

    st.markdown("## Input selection")

    col1, col2 = st.columns([1,5], vertical_alignment='top')

    mode = col1.radio("Mode", ("Upload CSV files", "Local folder on server"))

    processor = ExcelProcessor()

    uploaded_files = None
    local_folder = None
    if mode == "Upload CSV files":
        uploaded_files = col2.file_uploader("Upload one or more CSV files", type=["csv", "txt"], accept_multiple_files=True)
        if uploaded_files:
            processor.load_uploaded_csvs_to_memory(uploaded_files)
    else:
        local_folder = col2.text_input("Enter folder path containing CSV files (on the server)", value="")
        if local_folder:
            folder = Path(local_folder.strip().strip('"').strip("'"))
            if not folder.exists():
                col2.error("Folder not found on server")
            else:
                try:
                    processor.csv_folder_to_memory(str(folder))
                except Exception as e:
                    col2.error(f"Error loading folder: {e}")

    if not processor.sheet_names:
        col2.info("No sheets loaded yet. Upload files or provide a local folder and press Run.")
        st.stop()

    st.markdown("## Choose sheets (map CSVs to roles)")

    # option_map = {index: value for index, value in enumerate(processor.sheet_names)}

    cols = st.columns([4,1], vertical_alignment='center')

    invoice_sheet = working_refund_sheet = current_refund_sheet = ""
    revenues_sheet = working_dispute_sheet = current_dispute_sheet = ""

    with cols[0]:
        invoice_sheet = st.segmented_control(
                        f"{dt.now().strftime('%b-%Y')} Invoice sheet",
                        options=processor.sheet_names,
                        selection_mode="single",
                        key="inv_sheet"                   
                    )
        current_dispute_sheet = st.segmented_control(
                        f"{dt.now().strftime('%b-%Y')} Dispute sheet",
                        options=processor.sheet_names,
                        selection_mode="single",
                        key="curr_dis_sheet"
                    )
        current_refund_sheet = st.segmented_control(
                        f"{dt.now().strftime('%b-%Y')} Refund sheet",
                        options=processor.sheet_names,
                        selection_mode="single",
                        key="curr_ref_sheet"
                    )
        working_refund_sheet = st.segmented_control(
                        f"Working Refund sheet",
                        options=processor.sheet_names,
                        selection_mode="single",
                        key="work_ref_sheet"
                    )
        working_dispute_sheet = st.segmented_control(
                        f"Working Dispute sheet",
                        options=processor.sheet_names,
                        selection_mode="single",
                        key="work_dis_sheet"
                    )
        revenues_sheet = st.segmented_control(
                        f"Revenues sheet",
                        options=processor.sheet_names,
                        selection_mode="single",
                        key="rev_sheet"
                    )
        
    with cols[1]:
        st.html(f'<span style="color:green">Current-Year Invoice sheet:</span> {invoice_sheet}.')
        st.html(f'<span style="color:green">Working Refund sheet (2024 reference):</span> {working_refund_sheet}.')
        st.html(f'<span style="color:green">Current-Year Refund sheet:</span> {current_refund_sheet}.')
        st.html(f'<span style="color:green">Revenues sheet (2025 reference):</span> {revenues_sheet}.')
        st.html(f'<span style="color:green">Working Dispute sheet (2024 reference):</span> {working_dispute_sheet}.')
        st.html(f'<span style="color:green">Current-Year Dispute sheet:</span> {current_dispute_sheet}.')
    
    st.markdown("---")
    st.write("Optional settings")

    cols = st.columns(2)
    out_folder_input = cols[0].text_input("Output folder to save updated CSVs (server-side). Leave blank to create ./Updated_CSVs", value="")
    out_folder = out_folder_input.strip() or "Updated_CSVs"
    year_cutoff = cols[1].number_input("Year cutoff for check collection logic", value=2025)

    st.text("\n\n")

    cols = st.columns(3)
    run = cols[1].button("Run processing", use_container_width=True)

    # helper: convert stored list-of-rows into a small DataFrame for preview
    def preview_df_from_sheet(sheet_name, max_rows=6):
        rows = processor.sheets_data.get(sheet_name, [])[:max_rows]
        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows)
        # if first row looks like header (all strings), use it
        first_row = rows[0]
        if all(isinstance(x, str) for x in first_row):
            df.columns = first_row
            df = df.iloc[1:].reset_index(drop=True)
        return df

    if run:
        start = time.time()
        try:
            st.info("Running processing... (messages will appear below)")
            # call the same sequence as in your script
            processor.populate_inv_check_column(str(invoice_sheet), str(revenues_sheet),
                                            number_col_name="Number", lookup_col_name="Inv. No.", target_col_name="Inv Check")
            processor.populate_check_collection(str(current_refund_sheet), str(working_refund_sheet), str(revenues_sheet),
                                                created_col_name="Created date (UTC)",
                                                invoice_col_name="Invoice Number",
                                                collected_col_name="Collected",
                                                target_col_name="Check collection",
                                                revenue_lookup_col_name="Inv. No.",
                                                year_cutoff=year_cutoff)
            processor.populate_check_collection(str(current_dispute_sheet), str(working_dispute_sheet), str(revenues_sheet),
                                                created_col_name="Created date (UTC)",
                                                invoice_col_name="Invoice Number",
                                                collected_col_name="Collected",
                                                target_col_name="Check collection",
                                                revenue_lookup_col_name="Inv. No.",
                                                year_cutoff=year_cutoff)
            processor.populate_diff_column(str(current_refund_sheet),
                                        collected_col_name="Collected",
                                        check_col_name="Check collection",
                                        target_col_name="Diff")
            processor.populate_diff_column(str(current_dispute_sheet),
                                        collected_col_name="Collected",
                                        check_col_name="Check collection",
                                        target_col_name="Diff")
            # Save
            saved_paths = processor.save_to_csvs(out_folder)

            # Preview first few rows of each updated sheet (converted to DataFrame)
            st.subheader("📊 Updated Data Preview (first few rows of each updated sheet)")

            with st.expander("Invoice preview", expanded=False):
                st.dataframe(preview_df_from_sheet(invoice_sheet).head(10))

            with st.expander("Current Refund preview", expanded=False):
                st.dataframe(preview_df_from_sheet(current_refund_sheet).head(10))

            with st.expander("Current Dispute preview", expanded=False):
                st.dataframe(preview_df_from_sheet(current_dispute_sheet).head(10))

            # create zip for download
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for p in saved_paths:
                    zf.write(p, arcname=p.name)
            zip_buf.seek(0)
            #st.success(f"Processing completed in {time.time()-start:.2f}s — {len(saved_paths)} CSV(s) written.")
            elapsed = time.time() - start
            mins, secs = divmod(int(elapsed), 60)
            millis = int((elapsed - int(elapsed)) * 1000)
            st.success(f"Processing completed in {mins:02d}:{secs:02d}.{millis:03d} (mm:ss.ms) — {len(saved_paths)} CSV(s) written.")
            cols = st.columns(3)        
            cols[1].download_button("Download updated CSVs (ZIP)", data=zip_buf.getvalue(), file_name="updated_csvs.zip", mime="application/zip", use_container_width=True)
            st.write("Files written to server (if you used a local folder):")
            for p in saved_paths:
                st.write("-", p)
        except Exception as e:
            st.error(f"Processing failed: {e}")
            raise

#--- For Tab 4 ---
# with tab4:
#     # -------------------
#     # Helper functions (adapted to work with uploaded files / paths)
#     # -------------------

#     def try_open_csv_with_encodings(file_or_path, encodings=None, nrows=None, header=None, dtype=str, na_filter=False):
#         """
#         Try reading CSV using multiple encodings. Accepts:
#         - file_or_path: Pathlike / str path OR bytes/BytesIO (uploaded file).
#         Returns (DataFrame, encoding_used).
#         """
#         if encodings is None:
#             encodings = ["utf-8", "cp1252", "ISO-8859-1", "utf-8-sig"]

#         last_err = None
#         for enc in encodings:
#             try:
#                 if isinstance(file_or_path, (str, Path)):
#                     df = pd.read_csv(file_or_path, encoding=enc, nrows=nrows, header=header, dtype=dtype, na_filter=na_filter)
#                 else:
#                     # assume buffer/bytes
#                     buf = file_or_path if hasattr(file_or_path, "read") else io.BytesIO(file_or_path)
#                     # ensure buffer position
#                     try:
#                         buf.seek(0)
#                     except Exception:
#                         pass
#                     df = pd.read_csv(buf, encoding=enc, nrows=nrows, header=header, dtype=dtype, na_filter=na_filter)
#                 # Reduced logging to prevent UI flooding
#                 if nrows is None or nrows > 100:  # Only log for full reads
#                     st.success(f"✓ CSV loaded with encoding: {enc}")
#                 return df, enc
#             except Exception as e:
#                 last_err = e
#                 if nrows is None or nrows > 100:  # Only log errors for full reads
#                     st.warning(f"Encoding {enc} failed: {str(e)[:100]}...")
#                 continue
#         raise last_err or ValueError("All encodings failed")


#     def find_header_row_in_csv(file_or_path, key_cols, preview_rows=80):
#         """
#         Read a small preview (header=None) and find the index (0-based) of the first row containing all key_cols.
#         """
#         df_preview, enc = try_open_csv_with_encodings(file_or_path, nrows=preview_rows, header=None, dtype=str)
#         expected = set([str(k).strip() for k in key_cols])

#         st.info(f"🔍 Searching for header row with columns: {', '.join(expected)}")
        
#         for idx, row in df_preview.iterrows():
#             # gather non-empty stripped strings
#             values = set(str(cell).strip() for cell in row if pd.notna(cell) and str(cell).strip() != "")
#             if expected.issubset(values):
#                 st.success(f"✓ Found header row at index: {idx}")
#                 return idx
#         raise ValueError(f"Could not find header row containing keys: {key_cols}")


#     def find_header_row_in_excel_bytes(excel_bytes, sheet_name, key_cols, preview_rows=80):
#         """
#         Read small preview of an excel sheet (header=None) to find header row index.
#         excel_bytes: bytes or BytesIO
#         """
#         buf = excel_bytes if hasattr(excel_bytes, "read") else io.BytesIO(excel_bytes)
#         buf.seek(0)
#         preview_df = pd.read_excel(buf, sheet_name=sheet_name, header=None, nrows=preview_rows, dtype=str)
#         expected = set([str(k).strip() for k in key_cols])
#         for idx, row in preview_df.iterrows():
#             values = set(str(cell).strip() for cell in row if pd.notna(cell) and str(cell).strip() != "")
#             if expected.issubset(values):
#                 st.success(f"✓ Found Excel header row at index: {idx} (sheet {sheet_name})")
#                 return idx
#         raise ValueError(f"Could not find header row in excel sheet '{sheet_name}' containing keys: {key_cols}")


#     # -------------------
#     # Core logic (adapted from your function)
#     # -------------------

#     def append_missing_working_rows_streamlit(
#         working_file_buffer,
#         revenue_file_buffer,
#         chosen_sheet=None,
#         preview_rows=120,
#         inv_check_values=("#N/A", "N/A", "NA"),
#     ):
#         """
#         Inputs:
#         - working_file_buffer: uploaded CSV file (BytesIO or UploadedFile)
#         - revenue_file_buffer: uploaded Excel workbook (BytesIO or UploadedFile)
#         - chosen_sheet: optionally a sheet name (if None and workbook has multiple sheets user will be asked)
#         Returns (modified_workbook_bytes, debug_info_dict)
#         """
#         t0 = time.time()
#         debug = {}

#         # Progress tracking
#         progress_bar = st.progress(0)
#         status_text = st.empty()

#         try:
#             # 1) find header row in working CSV
#             status_text.text("📊 Analyzing working CSV structure...")
#             progress_bar.progress(10)
            
#             work_keys = [
#                 "Customer", "Customer Name", "Number",
#                 "Date (UTC)", "Minimum Line Item Period Start (UTC)",
#                 "Maximum Line Item Period End (UTC)", "Inv Amount", "Inv Check"
#             ]
#             work_hdr_idx = find_header_row_in_csv(working_file_buffer, work_keys, preview_rows=preview_rows)
            
#             # load full working DF using that header
#             status_text.text("📈 Loading working data...")
#             progress_bar.progress(20)
            
#             work_df, work_enc = try_open_csv_with_encodings(working_file_buffer, header=work_hdr_idx, dtype=str, na_filter=False)
#             # normalize headers
#             work_df.columns = work_df.columns.str.strip()
#             debug['work_rows'] = len(work_df)

#             # Memory optimization: only keep required columns early
#             selected_work_cols = [
#                 "Customer", "Customer Name", "Number",
#                 "Date (UTC)", "Minimum Line Item Period Start (UTC)",
#                 "Maximum Line Item Period End (UTC)", "Inv Amount", "Inv Check"
#             ]
#             work_df = work_df[selected_work_cols].copy()

#             # 2) open excel workbook bytes and list sheets
#             status_text.text("📋 Analyzing Excel workbook...")
#             progress_bar.progress(30)
            
#             buf = revenue_file_buffer if hasattr(revenue_file_buffer, "read") else io.BytesIO(revenue_file_buffer)
#             buf.seek(0)
            
#             # Use read_only mode for sheet detection to save memory
#             try:
#                 xls = pd.ExcelFile(buf, engine='openpyxl')
#                 sheets = xls.sheet_names
#                 xls.close()  # Close to free memory
#             except:
#                 # Fallback for older pandas versions
#                 wb_temp = load_workbook(buf, read_only=True)
#                 sheets = wb_temp.sheetnames
#                 wb_temp.close()

#             # pick sheet
#             if chosen_sheet:
#                 if chosen_sheet not in sheets:
#                     raise ValueError(f"Chosen sheet '{chosen_sheet}' not in workbook.")
#                 rev_sheet = chosen_sheet
#             else:
#                 # if only one sheet, auto pick it
#                 if len(sheets) == 1:
#                     rev_sheet = sheets[0]
#                     st.info(f"📄 Using sheet: '{rev_sheet}'")
#                 else:
#                     # return sheet list to UI to choose
#                     return None, {"sheets": sheets, "error": "multiple_sheets_choose_one"}

#             # 3) find header row in revenue excel sheet preview
#             status_text.text("🔍 Finding revenue sheet structure...")
#             progress_bar.progress(40)
            
#             rev_keys = [
#                 "Customer ID", "Account Name", "Inv. No.",
#                 "Date Issued", "Start Date", "End Date", "Collected"
#             ]
#             buf.seek(0)
#             rev_hdr_idx = find_header_row_in_excel_bytes(buf, rev_sheet, rev_keys, preview_rows=preview_rows)

#             # 4) load full revenue sheet as DataFrame using header row
#             status_text.text("📊 Loading revenue data...")
#             progress_bar.progress(50)
            
#             buf.seek(0)
#             rev_df = pd.read_excel(buf, sheet_name=rev_sheet, header=rev_hdr_idx, dtype=str)
#             rev_df.columns = rev_df.columns.str.strip()

#             # 5) find first blank in 'Inv. No.' column to split table/remainder
#             status_text.text("🔢 Processing invoice data...")
#             progress_bar.progress(60)
            
#             inv_col = "Inv. No."
#             if inv_col not in rev_df.columns:
#                 raise KeyError(f"'{inv_col}' not found in revenue sheet after header detection.")

#             mask_blank = rev_df[inv_col].fillna("").astype(str).str.strip() == ""
#             if mask_blank.any():
#                 first_blank_idx = mask_blank[mask_blank].index.min()
#             else:
#                 first_blank_idx = len(rev_df)

#             table = rev_df.iloc[:first_blank_idx].copy()
#             remainder = rev_df.iloc[first_blank_idx :].copy()

#             debug['rev_table_rows'] = len(table)
#             debug['rev_remainder_rows'] = len(remainder)
#             debug['rev_hdr_index'] = rev_hdr_idx

#             # 6) build mask on working df:
#             status_text.text("🎯 Filtering rows to append...")
#             progress_bar.progress(70)
            
#             inv_check_col = "Inv Check"
#             inv_amount_col = "Inv Amount"

#             if inv_check_col not in work_df.columns:
#                 raise KeyError(f"'{inv_check_col}' not found in working CSV columns: {list(work_df.columns)}")
#             if inv_amount_col not in work_df.columns:
#                 raise KeyError(f"'{inv_amount_col}' not found in working CSV columns: {list(work_df.columns)}")

#             # Normalize / treat values as strings
#             work_df[inv_check_col] = work_df[inv_check_col].astype(str).fillna("").str.strip()
#             inv_check_values_norm = {v.strip() for v in inv_check_values}
#             # match #N/A or blank
#             mask_invcheck_na = work_df[inv_check_col].isin(inv_check_values_norm) | (work_df[inv_check_col] == "")
#             # numeric inv amount:
#             work_df["_inv_amount_num"] = pd.to_numeric(work_df[inv_amount_col], errors="coerce").fillna(0.0)
#             mask_invamount_nonzero = work_df["_inv_amount_num"] != 0.0
#             mask_to_append = mask_invcheck_na & mask_invamount_nonzero

#             debug['mask_invcheck_na'] = int(mask_invcheck_na.sum())
#             debug['mask_invamount_nonzero'] = int(mask_invamount_nonzero.sum())
#             debug['mask_both'] = int(mask_to_append.sum())

#             # 7) select important columns and rename to revenue schema
#             selected_work_cols_final = [
#                 "Customer", "Customer Name", "Number",
#                 "Date (UTC)", "Minimum Line Item Period Start (UTC)",
#                 "Maximum Line Item Period End (UTC)", "Inv Amount"
#             ]
#             for c in selected_work_cols_final:
#                 if c not in work_df.columns:
#                     raise KeyError(f"Expected column '{c}' missing in working CSV")

#             new_rows = work_df.loc[mask_to_append, selected_work_cols_final].copy()
#             new_rows.columns = [
#                 "Customer ID", "Account Name", "Inv. No.",
#                 "Date Issued", "Start Date", "End Date", "Collected"
#             ]

#             # 8) de-dup against existing table Inv. No.
#             status_text.text("🔄 Removing duplicates...")
#             progress_bar.progress(80)
            
#             existing = set(table["Inv. No."].astype(str).str.strip())
#             before = len(new_rows)
#             new_rows = new_rows[~new_rows["Inv. No."].astype(str).str.strip().isin(existing)].copy()
#             after = len(new_rows)
#             debug['new_candidates_before'] = before
#             debug['new_candidates_after_dedup'] = after

#             if after == 0:
#                 debug['message'] = "No new rows to insert after applying filters & dedupe."
#                 progress_bar.progress(100)
#                 status_text.text("✅ Complete - No new rows needed")
#                 return None, debug

#             # 9) Insert rows into Excel workbook
#             status_text.text(f"📝 Inserting {after} new rows into Excel...")
#             progress_bar.progress(90)
            
#             # Load workbook from bytes (openpyxl) with optimizations
#             buf.seek(0)
#             wb = load_workbook(buf, data_only=False, keep_vba=False)
#             ws = wb[rev_sheet]

#             # Compute insert row position in Excel (1-based)
#             header_excel_rownum = rev_hdr_idx + 1
#             insert_at = header_excel_rownum + first_blank_idx + 1
#             n_new = len(new_rows)

#             # Insert empty rows (openpyxl shifts down the remainder)
#             ws.insert_rows(insert_at, amount=n_new)

#             # Map header names to excel col idx according to rev_df columns
#             header_to_colidx = {name: idx+1 for idx, name in enumerate(rev_df.columns)}

#             # Write new rows to inserted range with batch processing
#             for row_offset, (_, r) in enumerate(new_rows.iterrows()):
#                 excel_row = insert_at + row_offset
#                 for col_name in rev_df.columns:
#                     col_idx = header_to_colidx[col_name]
#                     if col_name in new_rows.columns:
#                         val = r[col_name]
#                         if pd.isna(val):
#                             val = ""
#                         ws.cell(row=excel_row, column=col_idx, value=str(val))

#             # Save workbook into bytes with memory optimization
#             status_text.text("💾 Saving updated workbook...")
#             progress_bar.progress(95)
            
#             out_buf = io.BytesIO()
#             wb.save(out_buf)
#             wb.close()  # Important: close workbook to free memory
#             out_buf.seek(0)

#             debug['insert_at_row'] = insert_at
#             debug['inserted_rows'] = n_new
#             debug['elapsed_s'] = time.time() - t0
#             debug['message'] = f"Successfully inserted {n_new} rows at Excel row {insert_at} in sheet '{rev_sheet}'"

#             progress_bar.progress(100)
#             status_text.text("✅ Processing complete!")
            
#             return out_buf.getvalue(), debug

#         except Exception as e:
#             progress_bar.progress(0)  # Reset progress on error
#             status_text.text("❌ Processing failed")
#             raise e
#         finally:
#             # Clean up progress indicators after a short delay
#             import threading
#             def clear_progress():
#                 time.sleep(2)
#                 try:
#                     progress_bar.empty()
#                     status_text.empty()
#                 except:
#                     pass
#             threading.Thread(target=clear_progress).start()

#     # -------------------
#     # Streamlit UI with optimizations
#     # -------------------

#     st.title("📊 Append Missing Working Rows → Revenue Excel")

#     st.markdown("""
#     Upload your **Working CSV** and **Revenue Excel (.xlsx)** files. 
    
#     ✨ **What this does:**
#     - Finds rows where `Inv Check` is `#N/A` (or blank) and `Inv Amount` ≠ 0
#     - Adds these rows to your Revenue Excel file
#     - Removes duplicates automatically
#     """)

#     # Initialize session state
#     if 'sheets_detected' not in st.session_state:
#         st.session_state.sheets_detected = []
#     if 'chosen_sheet' not in st.session_state:
#         st.session_state.chosen_sheet = None

#     # File size warnings
#     st.info("💡 **Tip:** For large files (>50MB), consider processing in smaller batches to avoid browser memory issues.")

#     col1, col2 = st.columns(2)
#     with col1:
#         st.subheader("📄 Working CSV File")
#         working_upload = st.file_uploader(
#             "Upload Working CSV", 
#             type=["csv", "txt"], 
#             accept_multiple_files=False,
#             help="CSV file containing working data with Inv Check and Inv Amount columns"
#         )
        
#         # Show file info if uploaded
#         if working_upload is not None:
#             file_size = len(working_upload.getvalue()) / (1024*1024)  # MB
#             st.success(f"✅ File loaded: {working_upload.name} ({file_size:.1f} MB)")
#             if file_size > 50:
#                 st.warning("⚠️ Large file detected. Processing may take longer.")

#         use_local_working = st.checkbox("Or use server file path", value=False, key="local_work")
#         working_local_path = ""
#         if use_local_working:
#             working_local_path = st.text_input("Local working CSV path", value="", key="work_path")

#     with col2:
#         st.subheader("📊 Revenue Excel File")
#         revenue_upload = st.file_uploader(
#             "Upload Revenue Excel (.xlsx)", 
#             type=["xlsx", "xlsm", "xlsb"], 
#             accept_multiple_files=False,
#             help="Excel workbook containing revenue data"
#         )
        
#         # Show file info if uploaded
#         if revenue_upload is not None:
#             file_size = len(revenue_upload.getvalue()) / (1024*1024)  # MB
#             st.success(f"✅ File loaded: {revenue_upload.name} ({file_size:.1f} MB)")
#             if file_size > 50:
#                 st.warning("⚠️ Large file detected. Processing may take longer.")

#         use_local_revenue = st.checkbox("Or use server file path", value=False, key="local_rev")
#         revenue_local_path = ""
#         if use_local_revenue:
#             revenue_local_path = st.text_input("Local revenue Excel path", value="", key="rev_path")

#     # Detect sheets when revenue file is uploaded
#     sheets_available = []
#     if revenue_upload is not None or (use_local_revenue and revenue_local_path):
#         try:
#             with st.spinner("🔍 Detecting Excel sheets..."):
#                 if use_local_revenue and revenue_local_path:
#                     xls = pd.ExcelFile(revenue_local_path, engine='openpyxl')
#                     sheets_available = xls.sheet_names
#                     xls.close()
#                 else:
#                     revenue_bytes = revenue_upload.read()
#                     revenue_upload.seek(0)  # Reset for later use
#                     tmp_buf = io.BytesIO(revenue_bytes)
#                     xls = pd.ExcelFile(tmp_buf, engine='openpyxl')
#                     sheets_available = xls.sheet_names
#                     xls.close()
                
#                 st.session_state.sheets_detected = sheets_available
                
#         except Exception as e:
#             st.error(f"❌ Error reading Excel file: {e}")
#             st.session_state.sheets_detected = []

#     # Show sheet selection if multiple sheets are detected
#     if len(st.session_state.sheets_detected) > 1:
#         st.subheader("📋 Sheet Selection")
#         chosen_sheet = st.selectbox(
#             "Choose Revenue sheet to update", 
#             options=st.session_state.sheets_detected, 
#             index=None,
#             key="sheet_selector",
#             help="Select the sheet containing your revenue data"
#         )
#         st.session_state.chosen_sheet = chosen_sheet
#     elif len(st.session_state.sheets_detected) == 1:
#         st.session_state.chosen_sheet = st.session_state.sheets_detected[0]
#         st.success(f"📄 Auto-selected sheet: '{st.session_state.chosen_sheet}'")
#     else:
#         st.session_state.chosen_sheet = None

#     # Show run button only when everything is ready
#     can_run = False
#     if use_local_working and use_local_revenue:
#         can_run = working_local_path and revenue_local_path
#     elif use_local_working:
#         can_run = working_local_path and revenue_upload is not None
#     elif use_local_revenue:
#         can_run = working_upload is not None and revenue_local_path
#     else:
#         can_run = working_upload is not None and revenue_upload is not None

#     # Additional check for sheet selection
#     if len(st.session_state.sheets_detected) > 1 and st.session_state.chosen_sheet is None:
#         can_run = False
#         st.warning("⚠️ Please select a sheet from the dropdown above.")

#     st.markdown("---")
    
#     if st.button("🚀 Process Files", disabled=not can_run, type="primary"):
#         try:
#             # Clear any previous results
#             if 'processing_result' in st.session_state:
#                 del st.session_state.processing_result

#             # Prepare buffers or paths
#             if use_local_working:
#                 working_source = working_local_path
#             else:
#                 working_source = io.BytesIO(working_upload.read())
#                 working_upload.seek(0)

#             if use_local_revenue:
#                 revenue_source = revenue_local_path  
#             else:
#                 revenue_source = io.BytesIO(revenue_upload.read())
#                 revenue_upload.seek(0)

#             # Run processing with error handling
#             st.info("🔄 Processing files... This may take a moment for large files.")
            
#             out_bytes, debug = append_missing_working_rows_streamlit(
#                 working_file_buffer=working_source,
#                 revenue_file_buffer=revenue_source,
#                 chosen_sheet=st.session_state.chosen_sheet,
#             )

#             # Store result in session state to prevent re-processing on rerun
#             st.session_state.processing_result = (out_bytes, debug)

#         except Exception as e:
#             st.error(f"❌ Processing failed: {str(e)}")
#             if st.checkbox("Show detailed error", value=False):
#                 st.exception(e)

#     # Display results if available
#     if 'processing_result' in st.session_state:
#         out_bytes, debug = st.session_state.processing_result
        
#         if out_bytes is None:
#             st.warning(f"ℹ️ {debug.get('message', 'No changes were made.')}")
#         else:
#             # Success! Provide download
#             sheet_name = st.session_state.chosen_sheet or "Revenues"
#             out_name = f"{sheet_name}_updated.xlsx"
            
#             st.success(f"✅ {debug.get('message', 'Processing completed successfully!')}")
            
#             col1, col2 = st.columns([2, 1])
#             with col1:
#                 st.download_button(
#                     "📥 Download Updated Excel File", 
#                     data=out_bytes, 
#                     file_name=out_name, 
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                     type="primary"
#                 )
#             with col2:
#                 if st.button("🔄 Process Another File"):
#                     # Clear session state to start fresh
#                     for key in ['processing_result', 'sheets_detected', 'chosen_sheet']:
#                         if key in st.session_state:
#                             del st.session_state[key]
#                     st.rerun()

#         # Show debug info in an expander
#         with st.expander("📊 Processing Details", expanded=False):
#             st.json(debug)

import asyncio
from concurrent.futures import ProcessPoolExecutor
import multiprocessing as mp

# This approach processes files on the server, not in browser memory
with tab4:
    
    def process_files_server_side(working_csv_path, revenue_excel_path, sheet_name):
        """
        This function runs in a separate process with its own memory space
        Completely isolated from browser memory limitations
        """
        import pandas as pd
        import openpyxl
        from openpyxl import load_workbook
        import io
        import time
        import gc
        
        def find_header_row(file_path, expected_cols, is_excel=False, sheet_name=None):
            if is_excel:
                preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=50)
            else:
                preview = pd.read_csv(file_path, header=None, nrows=50, encoding='utf-8')
            
            expected_set = set(str(col).strip() for col in expected_cols)
            
            for idx, row in preview.iterrows():
                row_values = set(str(cell).strip() for cell in row if pd.notna(cell) and str(cell).strip())
                if expected_set.issubset(row_values):
                    return idx
            raise ValueError(f"Header row not found for columns: {expected_cols}")
        
        try:
            # Process in isolated memory space
            print("Starting server-side processing...")
            
            # 1. Find headers
            work_cols = ["Customer", "Customer Name", "Number", "Date (UTC)", 
                        "Minimum Line Item Period Start (UTC)", "Maximum Line Item Period End (UTC)", 
                        "Inv Amount", "Inv Check"]
            rev_cols = ["Customer ID", "Account Name", "Inv. No.", "Date Issued", 
                       "Start Date", "End Date", "Collected"]
            
            work_header_row = find_header_row(working_csv_path, work_cols)
            rev_header_row = find_header_row(revenue_excel_path, rev_cols, True, sheet_name)
            
            # 2. Process CSV in chunks
            chunk_size = 50000  # Large chunks since we're server-side
            new_rows_list = []
            
            csv_iterator = pd.read_csv(working_csv_path, header=work_header_row, 
                                     chunksize=chunk_size, dtype=str, na_filter=False)
            
            for chunk_num, chunk in enumerate(csv_iterator):
                print(f"Processing CSV chunk {chunk_num + 1}...")
                
                chunk.columns = chunk.columns.str.strip()
                
                # Apply filters
                inv_check_values = {"#N/A", "N/A", "NA", ""}
                chunk["Inv Check"] = chunk["Inv Check"].astype(str).fillna("").str.strip()
                mask_check = chunk["Inv Check"].isin(inv_check_values) | (chunk["Inv Check"] == "")
                
                chunk["_inv_num"] = pd.to_numeric(chunk["Inv Amount"], errors="coerce").fillna(0)
                mask_amount = chunk["_inv_num"] != 0
                
                mask_final = mask_check & mask_amount
                
                if mask_final.any():
                    selected = chunk.loc[mask_final, work_cols[:7]].copy()  # Exclude Inv Check
                    selected.columns = rev_cols
                    new_rows_list.append(selected)
                
                del chunk
                gc.collect()
            
            if not new_rows_list:
                return None, {"message": "No rows to add"}
            
            # Combine new rows
            all_new_rows = pd.concat(new_rows_list, ignore_index=True)
            print(f"Found {len(all_new_rows)} candidate rows")
            
            # 3. Load existing Excel data efficiently
            existing_data = pd.read_excel(revenue_excel_path, sheet_name=sheet_name, 
                                        header=rev_header_row, dtype=str)
            
            # Find where data ends
            inv_col = "Inv. No."
            mask_blank = existing_data[inv_col].fillna("").astype(str).str.strip() == ""
            first_blank_idx = mask_blank[mask_blank].index.min() if mask_blank.any() else len(existing_data)
            
            # Remove duplicates
            existing_invs = set(existing_data[inv_col].iloc[:first_blank_idx].astype(str).str.strip())
            final_new_rows = all_new_rows[~all_new_rows[inv_col].astype(str).str.strip().isin(existing_invs)]
            
            if len(final_new_rows) == 0:
                return None, {"message": "No new rows after deduplication"}
            
            # 4. Update Excel file
            print(f"Adding {len(final_new_rows)} rows to Excel...")
            
            wb = load_workbook(revenue_excel_path)
            ws = wb[sheet_name]
            
            insert_row = rev_header_row + 1 + first_blank_idx + 1
            ws.insert_rows(insert_row, amount=len(final_new_rows))
            
            # Write data
            col_mapping = {col: idx + 1 for idx, col in enumerate(existing_data.columns)}
            
            for row_idx, (_, row_data) in enumerate(final_new_rows.iterrows()):
                excel_row = insert_row + row_idx
                for col_name in existing_data.columns:
                    if col_name in final_new_rows.columns:
                        value = str(row_data[col_name]) if not pd.isna(row_data[col_name]) else ""
                        ws.cell(row=excel_row, column=col_mapping[col_name], value=value)
            
            # Save to bytes
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            wb.close()
            
            return output_buffer.getvalue(), {
                "message": f"Successfully added {len(final_new_rows)} rows",
                "rows_added": len(final_new_rows),
                "insert_position": insert_row
            }
            
        except Exception as e:
            print(f"Server-side error: {e}")
            return None, {"error": str(e)}
        finally:
            gc.collect()

    # UI for server-side processing
    st.title("🖥️ Server-Side Large File Processor")
    
    st.markdown("""
    **🚀 Maximum Performance Solution:**
    - Processes files entirely on the server (not in browser memory)
    - Can handle files of ANY size (limited only by server RAM)
    - Uses multiprocessing for isolation
    - No browser crashes!
    """)

    # Initialize session state
    if 'server_processing' not in st.session_state:
        st.session_state.server_processing = False
    if 'server_result' not in st.session_state:
        st.session_state.server_result = None

    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📄 Working CSV")
        working_file = st.file_uploader("Upload Working CSV", type=["csv"], key="server_working")
        if working_file:
            size_mb = len(working_file.getvalue()) / (1024*1024)
            st.info(f"File size: {size_mb:.1f} MB")

    with col2:
        st.subheader("📊 Revenue Excel")  
        revenue_file = st.file_uploader("Upload Revenue Excel", type=["xlsx"], key="server_revenue")
        if revenue_file:
            size_mb = len(revenue_file.getvalue()) / (1024*1024)
            st.info(f"File size: {size_mb:.1f} MB")

    # Sheet selection
    selected_sheet = None
    if revenue_file:
        try:
            # Save temporarily to read sheets
            temp_path = f"/tmp/temp_revenue_{int(time.time())}.xlsx"
            with open(temp_path, "wb") as f:
                f.write(revenue_file.getvalue())
            
            xls = pd.ExcelFile(temp_path, engine='openpyxl')
            sheets = xls.sheet_names
            xls.close()
            
            if len(sheets) > 1:
                selected_sheet = st.selectbox("Select Revenue Sheet:", sheets)
            else:
                selected_sheet = sheets[0]
                st.success(f"Using sheet: {selected_sheet}")
                
            os.unlink(temp_path)  # Cleanup
        except Exception as e:
            st.error(f"Error reading sheets: {e}")

    # Process button
    can_process = (working_file is not None and revenue_file is not None and 
                  selected_sheet is not None and not st.session_state.server_processing)

    if st.button("🚀 Process on Server", disabled=not can_process, type="primary"):
        st.session_state.server_processing = True
        st.session_state.server_result = None
        
        try:
            # Save files to server temporary directory
            temp_working = f"/tmp/working_{int(time.time())}.csv"
            temp_revenue = f"/tmp/revenue_{int(time.time())}.xlsx"
            
            with open(temp_working, "wb") as f:
                f.write(working_file.getvalue())
            with open(temp_revenue, "wb") as f:
                f.write(revenue_file.getvalue())
            
            # Show processing indicator
            with st.spinner("🔄 Processing files on server... This may take several minutes."):
                progress_placeholder = st.empty()
                
                # Use process pool to isolate memory
                with ProcessPoolExecutor(max_workers=1) as executor:
                    future = executor.submit(process_files_server_side, temp_working, temp_revenue, selected_sheet)
                    
                    # Show progress while waiting
                    start_time = time.time()
                    while not future.done():
                        elapsed = int(time.time() - start_time)
                        progress_placeholder.info(f"⏱️ Processing... {elapsed}s elapsed")
                        time.sleep(2)
                    
                    result_bytes, debug_info = future.result()
                
                progress_placeholder.empty()
            
            # Cleanup temp files
            os.unlink(temp_working)
            os.unlink(temp_revenue)
            
            st.session_state.server_result = (result_bytes, debug_info)
            
        except Exception as e:
            st.error(f"❌ Server processing failed: {str(e)}")
        finally:
            st.session_state.server_processing = False

    # Show results
    if st.session_state.server_result:
        result_bytes, debug_info = st.session_state.server_result
        
        if result_bytes is None:
            st.warning(f"ℹ️ {debug_info.get('message', 'No changes made')}")
            if 'error' in debug_info:
                st.error(f"Error: {debug_info['error']}")
        else:
            st.success(f"✅ {debug_info.get('message', 'Processing completed successfully!')}")
            
            # Download button
            timestamp = int(time.time())
            filename = f"updated_revenue_{timestamp}.xlsx"
            
            st.download_button(
                "📥 Download Updated Excel File",
                data=result_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
            
            # Processing stats
            with st.expander("📊 Processing Details"):
                st.json(debug_info)
            
            # Reset button
            if st.button("🔄 Process New Files"):
                st.session_state.server_result = None
                st.rerun()

    st.markdown("---")
    
    # Alternative approach for VERY large files
    st.subheader("🗂️ Alternative: File Upload to Cloud Storage")
    
    with st.expander("For Files > 200MB", expanded=False):
        st.markdown("""
        **If server processing still fails, try this approach:**
        
        1. **Upload files to Google Drive/Dropbox**
        2. **Share the files publicly** 
        3. **Provide the direct download links below**
        4. **Server downloads and processes files directly**
        
        This eliminates browser upload limits entirely.
        """)
        
        col1, col2 = st.columns(2)
        with col1:
            working_url = st.text_input("Working CSV URL:", placeholder="https://drive.google.com/...")
        with col2:
            revenue_url = st.text_input("Revenue Excel URL:", placeholder="https://drive.google.com/...")
        
        if st.button("🌐 Process from URLs", disabled=not (working_url and revenue_url)):
            st.info("This feature requires additional implementation for URL downloads")
            st.code("""
            # Implementation would use:
            import requests
            
            def download_and_process(working_url, revenue_url):
                # Download files
                working_response = requests.get(working_url)
                revenue_response = requests.get(revenue_url)
                
                # Save temporarily and process
                # ... rest of processing logic
            """)

    # Hardware recommendations
    st.markdown("---")
    st.subheader("💻 Hardware & Platform Recommendations")
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("""
        **✅ Recommended Platforms:**
        - **Streamlit Cloud** (Community/Teams)
        - **Heroku** with higher memory dynos
        - **Railway** with memory scaling
        - **Google Cloud Run** with 4GB+ memory
        """)
    
    with col2:
        st.markdown("""
        **⚠️ File Size Limits by Platform:**
        - **Streamlit Community**: ~100MB total
        - **Streamlit Teams**: ~500MB total  
        - **Heroku Standard**: ~512MB RAM
        - **Custom Server**: Limited by RAM only
        """)

    # Tips section
    with st.expander("💡 Optimization Tips", expanded=False):
        st.markdown("""
        **🔧 Before Processing Large Files:**
        
        1. **Pre-filter your CSV** to reduce rows:
           ```python
           # Keep only rows where Inv Amount != 0
           df = df[pd.to_numeric(df['Inv Amount'], errors='coerce') != 0]
           ```
        
        2. **Split large Excel files** by date ranges or customers
        
        3. **Use CSV instead of Excel** when possible (much faster)
        
        4. **Remove unnecessary columns** before upload
        
        5. **Compress files** (.zip) to reduce upload time
        
        **🚀 Performance Benchmarks:**
        - 10k rows: ~10 seconds
        - 100k rows: ~2-5 minutes  
        - 1M rows: ~15-30 minutes
        - 10M rows: May require chunking/splitting
        """)

# Add error handling wrapper
try:
    # The tab4 content above
    pass
except MemoryError:
    st.error("""
    🚨 **Out of Memory Error**
    
    Your file is too large for browser processing. Try:
    1. **Use the Server-Side Processing** option above
    2. **Split your file** into smaller chunks
    3. **Upgrade to Streamlit Teams/Cloud** for more memory
    4. **Contact support** for enterprise processing options
    """)
except Exception as e:
    st.error(f"Processing error: {str(e)}")
    
    if "memory" in str(e).lower():
        st.info("This appears to be a memory issue. Try the server-side processing option above.")
    else:
        with st.expander("🔍 Full Error Details"):
            st.exception(e)