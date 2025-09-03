import os
import io
import time
import csv
import zipfile
import warnings
from pathlib import Path
import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import load_workbook
from pyexcelerate import Workbook
from datetime import datetime as dt
from collections import defaultdict
import threading
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
import re


# Page config and title
st.set_page_config(page_title="Financial Statement Processing", layout="wide")
st.title("💡 Automated Statement Processor")

# Define four tabs
tab1, tab2, tab3, tab4 = st.tabs([
    "📥 Import & Prepare Statements",
    "📂 Conert Excel Sheets to CSV",
    "⚙️ Complete Data Processing",
    "📈 Roll Over Revenue File Update"
])

#--- For Tab 1 ---
with tab1:
    st.set_page_config(page_title="Financial CSV Processor", layout="wide")


    def process_dataframes(df_invoice, df_refund, df_dispute, year_cutoff: int = 2025):
        """
        Vectorized processing adapted
        Returns updated (df_invoice, df_refund, df_dispute).
        """
        # Work on copies
        df_invoice = df_invoice.copy()
        df_refund = df_refund.copy()
        df_dispute = df_dispute.copy()

        # Remember original invoice column names to restore later
        orig_invoice_cols = list(df_invoice.columns)

        # Ensure required blank columns exist (do not overwrite existing)
        for col in ['Inv Check', 'Refund', 'Dispute', 'Inv Amount']:
            if col not in df_invoice.columns:
                df_invoice[col] = ''
        for col in ['Collected', 'Check collection', 'Diff']:
            if col not in df_refund.columns:
                df_refund[col] = ''
            if col not in df_dispute.columns:
                df_dispute[col] = ''

        # --- Refund sums ---
        if 'Invoice Number' in df_refund.columns:
            refund_key = 'Invoice Number'
        elif 'Inv. No.' in df_refund.columns:
            refund_key = 'Inv. No.'
        else:
            raise KeyError("Refund sheet missing invoice column ('Invoice Number' or 'Inv. No.')")

        if 'Amount Refunded' not in df_refund.columns:
            raise KeyError("Refund sheet missing 'Amount Refunded' column")

        refund_sums = df_refund.groupby(refund_key)['Amount Refunded'].sum()

        if 'Number' in df_invoice.columns:
            invoice_map_key = 'Number'
        elif 'Invoice Number' in df_invoice.columns:
            invoice_map_key = 'Invoice Number'
        else:
            raise KeyError("Invoice sheet missing invoice column ('Number' or 'Invoice Number')")

        df_invoice['Refund'] = df_invoice[invoice_map_key].map(refund_sums).fillna(0)

        # --- Dispute sums (status == 'lost') ---
        if 'Invoice Number' in df_dispute.columns:
            dis_key = 'Invoice Number'
        elif 'Inv. No.' in df_dispute.columns:
            dis_key = 'Inv. No.'
        else:
            raise KeyError("Dispute sheet missing invoice column ('Invoice Number' or 'Inv. No.')")

        if 'Dispute Status' not in df_dispute.columns or 'Disputed Amount' not in df_dispute.columns:
            raise KeyError("Dispute sheet missing 'Dispute Status' or 'Disputed Amount'")

        lost = df_dispute[df_dispute['Dispute Status'] == 'lost']
        dispute_sums = lost.groupby(dis_key)['Disputed Amount'].sum()
        df_invoice['Dispute'] = df_invoice[invoice_map_key].map(dispute_sums).fillna(0)

        # --- Inv Amount calculation ---
        for col in ['Amount Paid', 'Refund', 'Dispute']:
            if col not in df_invoice.columns:
                df_invoice[col] = 0
        df_invoice['Inv Amount'] = pd.to_numeric(df_invoice['Amount Paid'], errors='coerce').fillna(0) \
                                - pd.to_numeric(df_invoice['Refund'], errors='coerce').fillna(0) \
                                - pd.to_numeric(df_invoice['Dispute'], errors='coerce').fillna(0)

        # --- Collected in refund/dispute sheets ---
        if 'Amount' not in df_refund.columns or 'Amount Refunded' not in df_refund.columns:
            raise KeyError("Refund sheet must contain 'Amount' and 'Amount Refunded' for collected calculation")
        if 'Amount' not in df_dispute.columns or 'Disputed Amount' not in df_dispute.columns:
            raise KeyError("Dispute sheet must contain 'Amount' and 'Disputed Amount' for collected calculation")

        df_refund['Collected'] = pd.to_numeric(df_refund['Amount'], errors='coerce').fillna(0) \
                                - pd.to_numeric(df_refund['Amount Refunded'], errors='coerce').fillna(0)
        df_dispute['Collected'] = pd.to_numeric(df_dispute['Amount'], errors='coerce').fillna(0) \
                                - pd.to_numeric(df_dispute['Disputed Amount'], errors='coerce').fillna(0)

        # --- Make sure Check collection exists (placeholder 0) ---
        if 'Check collection' not in df_refund.columns:
            df_refund['Check collection'] = 0
        if 'Check collection' not in df_dispute.columns:
            df_dispute['Check collection'] = 0

        # restore Number column name if previously present (no destructive rename was done in this function)
        # Return updated DataFrames
        return df_invoice, df_refund, df_dispute


    def df_to_bytes_csv(df: pd.DataFrame, encoding: str = 'utf-8-sig') -> bytes:
        """Return CSV bytes for st.download_button."""
        with io.BytesIO() as buffer:
            df.to_csv(buffer, index=False, encoding=encoding)
            return buffer.getvalue()


    def main():
        st.title("Financial CSV Processor (Invoice / Refund / Dispute)")
        st.markdown(
            """
            Upload your three CSVs (Invoice, Refund, Dispute). The app will:
            * Add required columns if missing,
            * Compute Refund, Dispute, Inv Amount, Collected (vectorized),
            * Provide updated CSV downloads ( originals are not overwritten).
            """
        )

        upload_mode = st.radio("Input method", ("Upload files", "Provide file paths"))

        invoice_file = None
        refund_file = None
        dispute_file = None
        invoice_path_text = None
        refund_path_text = None
        dispute_path_text = None

        if upload_mode == "Upload files":
            invoice_file = st.file_uploader("Invoice CSV", type=["csv", "txt"], key="inv")
            refund_file  = st.file_uploader("Refund CSV",  type=["csv", "txt"], key="ref")
            dispute_file = st.file_uploader("Dispute CSV", type=["csv", "txt"], key="dis")
        else:
            invoice_path_text = st.text_input("Invoice CSV path (local)", "")
            refund_path_text  = st.text_input("Refund CSV path (local)", "")
            dispute_path_text = st.text_input("Dispute CSV path (local)", "")

        year_cutoff = st.number_input("Year cutoff (for check collection logic, unused placeholder)", value=2025, step=1)

        if st.button("Process files"):
            try:
                with st.spinner("Loading CSVs..."):
                    # Load according to chosen mode
                    if upload_mode == "Upload files":
                        if not (invoice_file and refund_file and dispute_file):
                            st.error("Please upload all three CSVs.")
                            return
                        df_invoice = pd.read_csv(invoice_file, dtype=str, encoding='utf-8-sig')
                        df_refund  = pd.read_csv(refund_file,  dtype=str, encoding='utf-8-sig')
                        df_dispute = pd.read_csv(dispute_file, dtype=str, encoding='utf-8-sig')
                        # set original path stems for naming
                        inv_stem = Path(invoice_file.name).stem
                        ref_stem = Path(refund_file.name).stem
                        dis_stem = Path(dispute_file.name).stem
                    else:
                        if not (invoice_path_text and refund_path_text and dispute_path_text):
                            st.error("Please provide all three file paths.")
                            return
                        # read from provided local paths
                        df_invoice = pd.read_csv(invoice_path_text, dtype=str, encoding='utf-8-sig')
                        df_refund  = pd.read_csv(refund_path_text,  dtype=str, encoding='utf-8-sig')
                        df_dispute = pd.read_csv(dispute_path_text, dtype=str, encoding='utf-8-sig')
                        inv_stem = Path(invoice_path_text).stem
                        ref_stem = Path(refund_path_text).stem
                        dis_stem = Path(dispute_path_text).stem

                # Process
                start = time.time()
                with st.spinner("Processing (vectorized)..."):
                    updated_invoice, updated_refund, updated_dispute = process_dataframes(
                        df_invoice, df_refund, df_dispute, year_cutoff=year_cutoff
                    )
                elapsed = time.time() - start
                mins, secs = divmod(int(elapsed), 60)
                millis = int((elapsed - int(elapsed)) * 1000)
                st.success(f"Processing completed in {mins:02d}:{secs:02d}.{millis:03d} (mm:ss.ms)")

                # Show small previews
                st.subheader("📊 Data Preview (first few rows of each sheet)")

                with st.expander("Invoice preview ", expanded=False):
                    st.dataframe(updated_invoice.head(10))

                with st.expander("Refund preview ", expanded=False):
                    st.dataframe(updated_refund.head(10))

                with st.expander("Dispute preview ", expanded=False):
                    st.dataframe(updated_dispute.head(10))

                # Prepare CSV bytes and provide download buttons
                inv_bytes = updated_invoice.to_csv(sep='\t', index=False, encoding='utf-8').encode('utf-8')
                ref_bytes = updated_refund.to_csv(sep='\t', index=False, encoding='utf-8').encode('utf-8')
                dis_bytes = updated_dispute.to_csv(sep='\t', index=False, encoding='utf-8').encode('utf-8')

                st.markdown("### Download updated Text files")

                # build a zip
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                    i = 0
                    name = [inv_stem, ref_stem, dis_stem]
                    for csv_byte in [inv_bytes, ref_bytes, dis_bytes]:
                        zf.writestr(f"{name[i]}_udpated.Txt", csv_byte)
                        i += 1
                zip_buf.seek(0)
                st.download_button("Download ZIP of all CSVs", data=zip_buf.getvalue(), file_name="Updated_Txt.zip", mime="application/zip")
                
                # individual buttons
                col1, col2, col3 = st.columns(3)
                col1.download_button(
                    label=f"Download {inv_stem}_updated.Txt",
                    data=inv_bytes,
                    file_name=f"{inv_stem}_updated.Txt",
                    mime="text/plain"
                )
                col2.download_button(
                    label=f"Download {ref_stem}_updated.Txt",
                    data=ref_bytes,
                    file_name=f"{ref_stem}_updated.Txt",
                    mime="text/plain"
                )
                col3.download_button(
                    label=f"Download {dis_stem}_updated.Txt",
                    data=dis_bytes,
                    file_name=f"{dis_stem}_updated.Txt",
                    mime="text/plain"
                )

            except Exception as exc:
                st.error(f"Error: {exc}")
                raise

    if __name__ == "__main__":
        main()

#--- For Tab 2 ---
with tab2:
    # Suppress openpyxl user warning about missing stylesheet (not harmful)
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    st.set_page_config(page_title="Excel -> CSV Converter", layout="wide")

    st.title("Excel -> CSV Converter")

    st.markdown(
        """
        Convert selected sheets from one or more Excel workbooks into CSVs.
        - For **local path** mode the app writes `All_CSVs` next to each workbook.
        - For **upload** mode the app provides downloadable CSV files (or a single ZIP).
        """
    )

    mode = st.radio("Input mode", ("Upload Excel files", "Local Excel paths (on this machine)"))

    def list_sheets_from_workbook_bytes(bytes_io):
        try:
            wb = load_workbook(bytes_io, read_only=True, data_only=True)
            return wb.sheetnames
        except Exception as e:
            raise RuntimeError(f"Failed reading workbook: {e}")

    def list_sheets_from_workbook_path(path: Path):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            return wb.sheetnames
        except Exception as e:
            raise RuntimeError(f"Failed reading workbook {path}: {e}")

    def convert_sheet_to_txt_bytes(wb_src, sheet_name: str) -> bytes:
        """Given an openpyxl Workbook object and a sheet name, return CSV bytes."""
        ws = wb_src[sheet_name]
        buf = io.StringIO()
        writer = csv.writer(buf, lineterminator="\n", delimiter='\t')
        for row in ws.iter_rows(values_only=True):
            # convert None -> "" so CSV cell isn't 'None'
            writer.writerow([("" if v is None else v) for v in row])
        return buf.getvalue().encode("utf-8-sig")

    def convert_uploaded_workbook_to_txt(uploaded_file, selected_sheets):
        """
        uploaded_file: streamlit UploadedFile
        selected_sheets: list of sheet names to convert
        returns list of tuples: (csv_filename, csv_bytes)
        """
        raw = uploaded_file.read()
        bio = io.BytesIO(raw)
        wb = load_workbook(bio, read_only=True, data_only=True)
        txt_files = []
        for sheet in selected_sheets:
            txt_bytes = convert_sheet_to_txt_bytes(wb, sheet)
            safe_sheet = "".join(c if (c.isalnum() or c in " _-") else "_" for c in sheet)[:50]
            txt_name = f"{Path(uploaded_file.name).stem}_{safe_sheet}.txt"
            txt_files.append((txt_name, txt_bytes))
        return txt_files

    def convert_local_workbook_to_txt(path: Path, selected_sheets, out_dir: Path):
        """
        path: Path to local workbook
        selected_sheets: list of sheet names to convert
        out_dir: directory to write CSVs to (will be created)
        returns list of written file paths
        """
        wb = load_workbook(path, read_only=True, data_only=True)
        out_dir.mkdir(parents=True, exist_ok=True)
        written = []
        for sheet in selected_sheets:
            txt_bytes = convert_sheet_to_txt_bytes(wb, sheet)
            safe_sheet = "".join(c if (c.isalnum() or c in " _-") else "_" for c in sheet)[:50]
            txt_path = out_dir / f"{path.stem}_{safe_sheet}.txt"
            with open(txt_path, "wb") as f:
                f.write(txt_bytes)
            written.append(txt_path)
        return written

    # UI: Upload or local paths
    workbooks_info = []  # list of dicts { 'name':..., 'type': 'upload'|'local', 'object':..., 'sheets':[...] , 'selected':[...] }

    if mode == "Upload Excel files":
        uploaded = st.file_uploader("Upload one or more Excel files (.xlsx/.xlsm) — use Shift/Ctrl to select multiple",
                                    type=["xlsx", "xlsm", "xls"], accept_multiple_files=True)
        if uploaded:
            st.info(f"{len(uploaded)} file(s) uploaded. Expand each to choose sheets.")
            for u in uploaded:
                try:
                    sheets = list_sheets_from_workbook_bytes(io.BytesIO(u.read()))
                    # after reading bytes above, need to reset pointer or re-create bytes for later conversion
                    u.seek(0)
                except Exception as exc:
                    st.error(f"Failed to read '{u.name}': {exc}")
                    sheets = []
                with st.expander(f"{u.name} ({len(sheets)} sheets)" , expanded=True):
                    st.write("Sheets:", sheets)
                    # default select all
                    selected = st.multiselect(f"Select sheets to convert for {u.name}", options=sheets, default=sheets, key=f"sel_{u.name}")
                    workbooks_info.append({"name": u.name, "type": "upload", "object": u, "sheets": sheets, "selected": selected})
        else:
            st.info("No uploads yet. Use the file uploader above.")

    else:  # Local path mode
        st.markdown("Enter full file paths (one per line). The app must run on the same machine that has these files.")
        paths_text = st.text_area("Paste Excel full paths here (one path per line):", height=120)
        if paths_text.strip():
            paths = [p.strip().strip('"').strip("'") for p in paths_text.splitlines() if p.strip()]
            for p in paths:
                path = Path(p)
                if not path.exists():
                    st.error(f"File not found: {p}")
                    continue
                try:
                    sheets = list_sheets_from_workbook_path(path)
                except Exception as exc:
                    st.error(f"Failed to read {p}: {exc}")
                    continue
                with st.expander(f"{path.name} ({len(sheets)} sheets)", expanded=True):
                    st.write("Sheets:", sheets)
                    selected = st.multiselect(f"Select sheets to convert for {path.name}", options=sheets, default=sheets, key=f"local_{path.name}")
                    workbooks_info.append({"name": str(path), "type": "local", "object": path, "sheets": sheets, "selected": selected})
        else:
            st.info("No local paths provided yet. Paste one or more paths above and press outside the text area to register.")

    # Convert action
    if workbooks_info:
        st.markdown("---")
        st.write("Selected workbooks & sheets to convert:")
        for info in workbooks_info:
            st.write(f"- **{info['name']}** → {len(info['selected'])} sheet(s) selected")
        if st.button("Convert selected sheets to Text Files"):
            start = time.time()
            all_downloadables = []  # for uploads: (csv_name, bytes)
            written_files = []      # for local: list of paths written
            errors = []
            with st.spinner("Converting..."):
                for info in workbooks_info:
                    try:
                        if info["type"] == "upload":
                            # uploaded file object was read earlier; ensure we're at start
                            uploaded_file = info["object"]
                            uploaded_file.seek(0)
                            txt_files = convert_uploaded_workbook_to_txt(uploaded_file, info["selected"])
                            all_downloadables.extend(txt_files)
                        else:
                            path = Path(info["object"])
                            out_dir = path.parent / "All_Texts"
                            written = convert_local_workbook_to_txt(path, info["selected"], out_dir)
                            written_files.extend(written)
                    except Exception as exc:
                        errors.append(f"{info['name']}: {exc}")

            elapsed = time.time() - start
            mins, secs = divmod(int(elapsed), 60)
            millis = int((elapsed - int(elapsed)) * 1000)
            st.success(f"Processing completed in {mins:02d}:{secs:02d}.{millis:03d} (mm:ss.ms)") 

            if errors:
                st.error("Some files failed:")
                for e in errors:
                    st.write(f"- {e}")

            if written_files:
                st.success(f"Wrote {len(written_files)} Txt files in local 'All_Txts' folders.")
                st.write("Files written:")
                for p in written_files:
                    st.write(f"- {p}")

            if all_downloadables:
                st.success(f"Prepared {len(all_downloadables)} CSV download(s) from uploads.")
                # If many CSVs, offer a ZIP download; otherwise show individual buttons
                if len(all_downloadables) == 1:
                    name, b = all_downloadables[0]
                    st.download_button("Download CSV", data=b, file_name=name, mime="text/csv")
                else:
                    # build a zip
                    zip_buf = io.BytesIO()
                    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                        for name, b in all_downloadables:
                            zf.writestr(name, b)
                    zip_buf.seek(0)
                    st.download_button("Download ZIP of all Txts", data=zip_buf.getvalue(), file_name="converted_txts.zip", mime="application/zip")

            st.success(f"Conversion completed in {elapsed:.2f} seconds.")
            st.balloons()

    else:
        st.info("No workbooks selected yet. Upload files or provide local paths and choose sheets.")

    st.markdown("---")
    st.caption("This tool uses openpyxl in read_only mode and csv.writer, similar to your original script.")

#--- For Tab 3 ---
with tab3:
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    st.set_page_config(page_title="CSV Processor (ExcelProcessor)", layout="wide")
    st.title("CSV Processor (wraps ExcelProcessor)")

    # ---------------------------
    # Paste / recreate ExcelProcessor here
    # ---------------------------
    class ExcelProcessor:
        def __init__(self):
            self.sheets_data = {}  # sheet_name -> list of rows
            self.sheet_names = []

        def csv_folder_to_memory(self, csv_folder: str) -> None:
            start = time.time()
            folder = Path(csv_folder)
            csvs = sorted(folder.glob("*.csv"))
            if not csvs:
                raise RuntimeError(f"No CSV files found in {csv_folder!r}")
            for csv_path in csvs:
                with open(csv_path, newline='', encoding='utf-8') as f:
                    reader = list(csv.reader(f))
                    sheet_data = [row for row in reader]
                sheet_name = csv_path.stem[:31]
                self.sheets_data[sheet_name] = sheet_data
                self.sheet_names.append(sheet_name)
                st.write(f"Loaded {csv_path.name} as '{sheet_name}'")
            elapsed = time.time() - start
            st.write(f"Loaded {len(csvs)} CSVs into memory in {elapsed:.2f}s")

        def load_uploaded_csvs_to_memory(self, uploaded_files):
            """Handle both CSV and tab-delimited TXT files."""
            start = time.time()
            for up in uploaded_files:
                raw = up.read()
                try:
                    text = raw.decode("utf-8")
                except Exception:
                    text = raw.decode("utf-8", errors="replace")
                    
                # Detect if file is tab-delimited
                is_tab_delimited = '\t' in text.split('\n')[0]
                
                if is_tab_delimited:
                    # Handle as tab-delimited
                    reader = list(csv.reader(io.StringIO(text), delimiter='\t'))
                else:
                    # Handle as regular CSV
                    reader = list(csv.reader(io.StringIO(text)))
                    
                sheet_name = Path(up.name).stem[:31]
                self.sheets_data[sheet_name] = reader
                self.sheet_names.append(sheet_name)
                st.write(f"Loaded (upload) {up.name} as '{sheet_name}'")
                up.seek(0)
            elapsed = time.time() - start
            st.write(f"Loaded {len(uploaded_files)} uploaded files in {elapsed:.2f}s")

        def find_header_row(self, sheet_name: str, key_cols: list, preview_rows: int = 50) -> int:
            if sheet_name not in self.sheets_data:
                raise ValueError(f"Sheet {sheet_name!r} not found")
            expected = set(key_cols)
            sheet_data = self.sheets_data[sheet_name]
            for idx, row in enumerate(sheet_data[:preview_rows]):
                #Handle both list and string inputs (in case of tab-delimeted text)
                if isinstance(row, str):
                    #split by tab if it's a string (from txt file)
                    values = [str(v).strip() for v in row.split('\t') if v]
                else:
                    #Handle as list (from csv file)
                    values = [str(v).strip() for v in row if v is not None and str(v).strip() != ""]

                #create set of non-empty values
                values_set = set(values)

                #Debug output (optional) 
                #st.write(f"Row {idx}: Found Headers: {values_set}")

                if expected.issubset(values_set):
                    st.success(f"Found header row at index {idx} in sheet {sheet_name}")
                    return idx

            #If we get here, no header was found
            st.error(f"Could not find header row with columns {key_cols} in first {preview_rows} rows of sheet {sheet_name}")
            raise ValueError(f"Could not find header {key_cols} in sheet {sheet_name!r}")

        def choose_sheet(self, prompt: str) -> str:
            print(f"\n{prompt}")
            for i, name in enumerate(self.sheet_names, 1):
                print(f"  {i}. {name}")
            choice = int(input(f"Select [1–{len(self.sheet_names)}]: "))
            return self.sheet_names[choice - 1]

        def populate_inv_check_column(self, invoice_sheet: str, revenue_sheet: str,
                                    number_col_name: str = "Number",
                                    lookup_col_name: str = "Inv. No.",
                                    target_col_name: str = "Inv Check") -> None:
            start_time = time.time()
            inv_hdr = self.find_header_row(invoice_sheet, [number_col_name, target_col_name])
            rev_hdr = self.find_header_row(revenue_sheet, [lookup_col_name])
            inv_data = self.sheets_data[invoice_sheet]
            rev_data = self.sheets_data[revenue_sheet]
            rev_header = [col.strip() if col else col for col in rev_data[rev_hdr]]
            lookup_col_idx = rev_header.index(lookup_col_name)
            lookup_set = set()
            for row in rev_data[rev_hdr + 1:]:
                if lookup_col_idx < len(row) and row[lookup_col_idx]:
                    lookup_set.add(str(row[lookup_col_idx]).strip())
            inv_header = [col.strip() if col else col for col in inv_data[inv_hdr]]
            number_col_idx = inv_header.index(number_col_name)
            target_col_idx = inv_header.index(target_col_name)
            for i, row in enumerate(inv_data[inv_hdr + 1:], start=inv_hdr + 1):
                if number_col_idx < len(row) and row[number_col_idx]:
                    number_val = str(row[number_col_idx]).strip()
                    result = number_val if number_val in lookup_set else "#N/A"
                    new_row = list(row)
                    while len(new_row) <= target_col_idx:
                        new_row.append("")
                    new_row[target_col_idx] = result
                    inv_data[i] = new_row
            elapsed = time.time() - start_time
            st.write(f"Populated '{target_col_name}' on '{invoice_sheet}' in {elapsed:.2f}s")

        def populate_check_collection(self, current_sheet: str, working_sheet: str, revenues_sheet: str,
                                    created_col_name: str = "Created date (UTC)",
                                    invoice_col_name: str = "Invoice Number",
                                    collected_col_name: str = "Collected",
                                    target_col_name: str = "Check collection",
                                    revenue_lookup_col_name: str = "Inv. No.",
                                    year_cutoff: int = 2025) -> None:
            start = time.time()
            current_hdr = self.find_header_row(current_sheet, [created_col_name, invoice_col_name, target_col_name])
            working_hdr = self.find_header_row(working_sheet, [invoice_col_name, collected_col_name])
            revenue_hdr = self.find_header_row(revenues_sheet, [revenue_lookup_col_name, collected_col_name])
            current_data = self.sheets_data[current_sheet]
            working_data = self.sheets_data[working_sheet]
            revenue_data = self.sheets_data[revenues_sheet]
            def build_lookup_sums(data, header_idx, invoice_col, collected_col):
                header = data[header_idx]
                inv_idx = header.index(invoice_col)
                col_idx = header.index(collected_col)
                sums = defaultdict(float)
                for row in data[header_idx + 1:]:
                    if (inv_idx < len(row) and col_idx < len(row) and row[inv_idx] and row[col_idx]):
                        try:
                            sums[row[inv_idx]] += float(row[col_idx])
                        except (ValueError, TypeError):
                            pass
                return sums
            working_sums = build_lookup_sums(working_data, working_hdr, invoice_col_name, collected_col_name)
            revenue_sums = build_lookup_sums(revenue_data, revenue_hdr, revenue_lookup_col_name, collected_col_name)
            header = current_data[current_hdr]
            created_idx = header.index(created_col_name)
            invoice_idx = header.index(invoice_col_name)
            target_idx = header.index(target_col_name)
            for i, row in enumerate(current_data[current_hdr + 1:], start=current_hdr + 1):
                if (created_idx < len(row) and invoice_idx < len(row) and row[created_idx] and row[invoice_idx]):
                    created = row[created_idx]
                    if isinstance(created, dt):
                        year = created.year
                    else:
                        try:
                            year = int(str(created)[:4])
                        except (ValueError, TypeError):
                            year = 0
                    invoice_num = row[invoice_idx]
                    total = (working_sums.get(invoice_num, 0.0) if year < year_cutoff else revenue_sums.get(invoice_num, 0.0))
                    new_row = list(row)
                    while len(new_row) <= target_idx:
                        new_row.append("")
                    new_row[target_idx] = total
                    current_data[i] = new_row
            elapsed = time.time() - start
            st.write(f"Populated '{target_col_name}' on '{current_sheet}' in {elapsed:.2f}s")

        def populate_diff_column(self, sheet_name: str,
                            collected_col_name: str = "Collected",
                            check_col_name: str = "Check collection",
                            target_col_name: str = "Diff") -> None:
            start = time.time()
            header_idx = self.find_header_row(sheet_name, [collected_col_name, check_col_name, target_col_name])
            data = self.sheets_data[sheet_name]
            header = data[header_idx]
            collected_idx = header.index(collected_col_name)
            check_idx = header.index(check_col_name)
            target_idx = header.index(target_col_name)
            for i, row in enumerate(data[header_idx + 1:], start=header_idx + 1):
                collected = 0.0
                checked = 0.0
                if collected_idx < len(row) and row[collected_idx]:
                    try:
                        collected = float(row[collected_idx])
                    except (ValueError, TypeError):
                        pass
                if check_idx < len(row) and row[check_idx]:
                    try:
                        checked = float(row[check_idx])
                    except (ValueError, TypeError):
                        pass
                diff = collected - checked
                new_row = list(row)
                while len(new_row) <= target_idx:
                    new_row.append("")
                new_row[target_idx] = diff
                data[i] = new_row
            elapsed = time.time() - start
            st.write(f"Populated '{target_col_name}' on '{sheet_name}' in {elapsed:.2f}s")

        def save_to_csvs(self, out_folder: str = "All_TXTs", encoding: str = "utf-8-sig") -> list[Path]:
            start = time.time()
            out_dir = Path(out_folder)
            out_dir.mkdir(parents=True, exist_ok=True)
            written = []
            
            # Map of processed sheets to their safe names
            processed_sheets = {
                invoice_sheet: "Invoice",
                current_refund_sheet: "Refund",
                current_dispute_sheet: "Dispute"
            }
            
            for sheet_name, sheet_type in processed_sheets.items():
                if sheet_name in self.sheets_data:
                    safe = "".join(c if (c.isalnum() or c in " _-") else "_" for c in sheet_name)[:50]
                    txt_path = out_dir / f"{safe}_updated.txt"
                    
                    with open(txt_path, "w", newline="", encoding=encoding) as f:
                        writer = csv.writer(f, delimiter='\t')  # Use tab delimiter
                        rows = self.sheets_data.get(sheet_name, [])
                        for row in rows:
                            row_out = [("" if cell is None else str(cell)) for cell in row]
                            writer.writerow(row_out)
                    written.append(txt_path)
                    st.write(f"Saved {sheet_type} data as {txt_path}")

            elapsed = time.time() - start
            st.write(f"Saved {len(written)} CSVs to '{out_dir}' in {elapsed:.2f}s")
            return written

    st.markdown("## Input selection")

    col1, col2 = st.columns([1,5], vertical_alignment='top')

    mode = col1.radio("Mode", ("Upload CSV files", "Local folder on server"))

    processor = ExcelProcessor()

    uploaded_files = None
    local_folder = None
    if mode == "Upload CSV files":
        uploaded_files = col2.file_uploader("Upload one or more CSV files", type=["csv", "txt"], accept_multiple_files=True)
        if uploaded_files:
            processor.load_uploaded_csvs_to_memory(uploaded_files)
    else:
        local_folder = col2.text_input("Enter folder path containing CSV files (on the server)", value="")
        if local_folder:
            folder = Path(local_folder.strip().strip('"').strip("'"))
            if not folder.exists():
                col2.error("Folder not found on server")
            else:
                try:
                    processor.csv_folder_to_memory(str(folder))
                except Exception as e:
                    col2.error(f"Error loading folder: {e}")

    if not processor.sheet_names:
        col2.info("No sheets loaded yet. Upload files or provide a local folder and press Run.")
        st.stop()

    st.markdown("## Choose sheets (map CSVs to roles)")

    # option_map = {index: value for index, value in enumerate(processor.sheet_names)}

    cols = st.columns([4,1], vertical_alignment='center')

    invoice_sheet = working_refund_sheet = current_refund_sheet = ""
    revenues_sheet = working_dispute_sheet = current_dispute_sheet = ""

    with cols[0]:
        invoice_sheet = st.segmented_control(
                        f"{dt.now().strftime('%b-%Y')} - Invoice sheet",
                        options=processor.sheet_names,
                        selection_mode="single",
                        key="inv_sheet"                   
                    )
        current_dispute_sheet = st.segmented_control(
                        f"{dt.now().strftime('%b-%Y')} - Dispute sheet",
                        options=processor.sheet_names,
                        selection_mode="single",
                        key="curr_dis_sheet"
                    )
        current_refund_sheet = st.segmented_control(
                        f"{dt.now().strftime('%b-%Y')} - Refund sheet",
                        options=processor.sheet_names,
                        selection_mode="single",
                        key="curr_ref_sheet"
                    )
        working_refund_sheet = st.segmented_control(
                        f"Working Refund sheet",
                        options=processor.sheet_names,
                        selection_mode="single",
                        key="work_ref_sheet"
                    )
        working_dispute_sheet = st.segmented_control(
                        f"Working Dispute sheet",
                        options=processor.sheet_names,
                        selection_mode="single",
                        key="work_dis_sheet"
                    )
        revenues_sheet = st.segmented_control(
                        f"Revenues sheet",
                        options=processor.sheet_names,
                        selection_mode="single",
                        key="rev_sheet"
                    )
        
    with cols[1]:
        st.html(f'{dt.now().strftime('%b-%Y')}<span style="color:green"> - Invoice sheet:</span> {invoice_sheet}.')
        st.html(f'{dt.now().strftime('%b-%Y')}<span style="color:green"> - Dispute sheet:</span> {current_dispute_sheet}.')
        st.html(f'{dt.now().strftime('%b-%Y')}<span style="color:green"> - Refund sheet:</span> {current_refund_sheet}.')
        st.html(f'<span style="color:green">Working Refund sheet:</span> {working_refund_sheet}.')
        st.html(f'<span style="color:green">Working Dispute sheet:</span> {working_dispute_sheet}.')
        st.html(f'<span style="color:green">Revenues sheet:</span> {revenues_sheet}.')
    
    st.markdown("---")
    st.write("Optional settings")

    cols = st.columns(2)
    out_folder_input = cols[0].text_input("Output folder to save updated CSVs (server-side). Leave blank to create ./Updated_CSVs", value="")
    out_folder = out_folder_input.strip() or "Updated_CSVs"
    year_cutoff = cols[1].number_input("Year cutoff for check collection logic", value=2025)

    st.text("\n\n")

    cols = st.columns(3)
    run = cols[1].button("Run processing", use_container_width=True)

    # helper: convert stored list-of-rows into a small DataFrame for preview
    def preview_df_from_sheet(sheet_name, max_rows=6):
        rows = processor.sheets_data.get(sheet_name, [])[:max_rows]
        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows)
        # if first row looks like header (all strings), use it
        first_row = rows[0]
        if all(isinstance(x, str) for x in first_row):
            df.columns = first_row
            df = df.iloc[1:].reset_index(drop=True)
        return df

    if run:
        start = time.time()
        try:
            st.info("Running processing... (messages will appear below)")
            # call the same sequence as in your script
            processor.populate_inv_check_column(str(invoice_sheet), str(revenues_sheet),
                                            number_col_name="Number", lookup_col_name="Inv. No.", target_col_name="Inv Check")
            processor.populate_check_collection(str(current_refund_sheet), str(working_refund_sheet), str(revenues_sheet),
                                                created_col_name="Created date (UTC)",
                                                invoice_col_name="Invoice Number",
                                                collected_col_name="Collected",
                                                target_col_name="Check collection",
                                                revenue_lookup_col_name="Inv. No.",
                                                year_cutoff=year_cutoff)
            processor.populate_check_collection(str(current_dispute_sheet), str(working_dispute_sheet), str(revenues_sheet),
                                                created_col_name="Created date (UTC)",
                                                invoice_col_name="Invoice Number",
                                                collected_col_name="Collected",
                                                target_col_name="Check collection",
                                                revenue_lookup_col_name="Inv. No.",
                                                year_cutoff=year_cutoff)
            processor.populate_diff_column(str(current_refund_sheet),
                                        collected_col_name="Collected",
                                        check_col_name="Check collection",
                                        target_col_name="Diff")
            processor.populate_diff_column(str(current_dispute_sheet),
                                        collected_col_name="Collected",
                                        check_col_name="Check collection",
                                        target_col_name="Diff")
            # Save
            saved_paths = processor.save_to_csvs(out_folder)

            # Preview first few rows of each updated sheet (converted to DataFrame)
            st.subheader("📊 Updated Data Preview (first few rows of each updated sheet)")

            with st.expander(f"{dt.now().strftime('%b-%Y')} - Invoice preview", expanded=False):
                st.dataframe(preview_df_from_sheet(invoice_sheet).head(10))

            with st.expander(f"{dt.now().strftime('%b-%Y')} - Refund preview", expanded=False):
                st.dataframe(preview_df_from_sheet(current_refund_sheet).head(10))

            with st.expander(f"{dt.now().strftime('%b-%Y')} - Dispute preview", expanded=False):
                st.dataframe(preview_df_from_sheet(current_dispute_sheet).head(10))

            # create zip for download
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for p in saved_paths:
                    zf.write(p, arcname=p.name)
            zip_buf.seek(0)
            elapsed = time.time() - start
            mins, secs = divmod(int(elapsed), 60)
            millis = int((elapsed - int(elapsed)) * 1000)
            st.success(f"Processing completed in {mins:02d}:{secs:02d}.{millis:03d} (mm:ss.ms) — {len(saved_paths)} CSV(s) written.")
            cols = st.columns(3)        
            cols[1].download_button("Download processed TXTs (ZIP)", data=zip_buf.getvalue(), file_name="processed_txts.zip", mime="application/zip", use_container_width=True)
            st.write("Files written to server (if you used a local folder):")
            for p in saved_paths:
                st.write("-", p)
        except Exception as e:
            st.error(f"Processing failed: {e}")
            raise

#--- For Tab 4 ---
with tab4:
    # -------------------
    # Helper functions
    # -------------------
    def try_open_csv_with_encodings(file_or_path, encodings=None, nrows=None, header=None, dtype=str, na_filter=False):
        """
        Try reading CSV/TXT using multiple encodings. Handles both CSV and tab-delimited files.
        """
        if encodings is None:
            encodings = ["utf-8", "cp1252", "ISO-8859-1", "utf-8-sig"]

        last_err = None
        for enc in encodings:
            try:
                if isinstance(file_or_path, (str, Path)):
                    # Try reading as tab-delimited first
                    try:
                        df = pd.read_csv(file_or_path, encoding=enc, nrows=nrows, header=header, 
                                    dtype=dtype, na_filter=na_filter, sep='\t')
                    except:
                        # If that fails, try regular CSV
                        df = pd.read_csv(file_or_path, encoding=enc, nrows=nrows, header=header, 
                                    dtype=dtype, na_filter=na_filter)
                else:
                    # Handle uploaded file
                    buf = file_or_path if hasattr(file_or_path, "read") else io.BytesIO(file_or_path)
                    try:
                        buf.seek(0)
                    except Exception:
                        pass
                    # Try reading as tab-delimited first
                    try:
                        df = pd.read_csv(buf, encoding=enc, nrows=nrows, header=header, 
                                    dtype=dtype, na_filter=na_filter, sep='\t')
                    except:
                        buf.seek(0)
                        df = pd.read_csv(buf, encoding=enc, nrows=nrows, header=header, 
                                    dtype=dtype, na_filter=na_filter)
                        
                if nrows is None or nrows > 100:
                    st.success(f"✓ File loaded with encoding: {enc}")
                return df, enc
                
            except Exception as e:
                last_err = e
                if nrows is None or nrows > 100:
                    st.warning(f"Encoding {enc} failed: {str(e)[:120]}...")
                continue
                
        raise last_err or ValueError("All encodings failed")


    def find_header_row_in_csv(file_or_path, key_cols, preview_rows=80):
        """
        Read a small preview (header=None) and find the index (0-based) of the first row containing all key_cols.
        """
        df_preview, enc = try_open_csv_with_encodings(file_or_path, nrows=preview_rows, header=None, dtype=str)
        expected = set([str(k).strip() for k in key_cols])

        st.info(f"🔍 Searching for header row with columns: {', '.join(expected)}")

        for idx, row in df_preview.iterrows():
            # Handle both list and string inputs (in case of tab-delimited text)
            if isinstance(row, str):
                # Split by tab if it's a string (from txt file)
                values = [str(v).strip() for v in row.split('\t') if v]
            else:
                # Handle as list (from csv file)
                values = [str(v).strip() for v in row if pd.notna(v) and str(v).strip() != ""]
            
            # Create set of non-empty values
            values_set = set(values)
            
            # Debug output (uncomment if needed)
            # st.write(f"Row {idx}: Found Headers: {values_set}")
            
            if expected.issubset(values):
                st.success(f"✓ Found header row at index: {idx}")
                return idx
            
        raise ValueError(f"Could not find header row containing keys: {key_cols}")


    def find_header_row_in_excel_bytes(excel_bytes, sheet_name, key_cols, preview_rows=80):
        """
        Read small preview of an excel sheet (header=None) to find header row index.
        excel_bytes: bytes or BytesIO
        """
        buf = excel_bytes if hasattr(excel_bytes, "read") else io.BytesIO(excel_bytes)
        buf.seek(0)
        preview_df = pd.read_excel(buf, sheet_name=sheet_name, header=None, nrows=preview_rows, dtype=str)
        expected = set([str(k).strip() for k in key_cols])
        for idx, row in preview_df.iterrows():
            values = set(str(cell).strip() for cell in row if pd.notna(cell) and str(cell).strip() != "")
            if expected.issubset(values):
                st.success(f"✓ Found Excel header row at index: {idx} (sheet {sheet_name})")
                return idx
        raise ValueError(f"Could not find header row in excel sheet '{sheet_name}' containing keys: {key_cols}")


    # -------------------
    # Core logic (append function)
    # -------------------
    def append_missing_working_rows_streamlit(
        working_file_buffer,
        revenue_file_buffer,
        chosen_sheet=None,
        preview_rows=120,
        inv_check_values=("#N/A", "N/A", "NA"),
    ):
        """
        Inputs:
        - working_file_buffer: uploaded CSV file (BytesIO or UploadedFile) OR local path string
        - revenue_file_buffer: uploaded Excel workbook (BytesIO or UploadedFile) OR local path string
        - chosen_sheet: optionally a sheet name (if None and workbook has multiple sheets user will be asked)
        Returns (modified_workbook_bytes, debug_info_dict)
        """
        t0 = time.time()
        debug = {}

        # Progress placeholders
        progress_bar = st.progress(0)
        status_text = st.empty()
        elapsed_placeholder = st.empty()

        # Thread control event to stop the elapsed-updater
        _elapsed_stop = threading.Event()

        def _elapsed_updater():
            # Update once immediately then every 1 second
            while not _elapsed_stop.is_set():
                elapsed = time.time() - t0
                mins, secs = divmod(int(elapsed), 60)
                millis = int((elapsed - int(elapsed)) * 1000)
                elapsed_placeholder.info(f"⏱️ Elapsed {mins:02d}:{secs:02d}.{millis:03d}")
                # wait with event so it's interruptible
                _elapsed_stop.wait(1.0)

        # start elapsed updater thread
        threading.Thread(target=_elapsed_updater, daemon=True).start()

        try:
            # 1) find header row in working CSV
            status_text.text("📊 Analyzing working CSV structure...")
            progress_bar.progress(10)

            work_keys = [
                "Customer", "Customer Name", "Number",
                "Date (UTC)", "Minimum Line Item Period Start (UTC)",
                "Maximum Line Item Period End (UTC)", "Inv Amount", "Inv Check"
            ]
            work_hdr_idx = find_header_row_in_csv(working_file_buffer, work_keys, preview_rows=preview_rows)

            # 2) load full working DF using that header
            status_text.text("📈 Loading working data...")
            progress_bar.progress(20)
            work_df, work_enc = try_open_csv_with_encodings(working_file_buffer, header=work_hdr_idx, dtype=str, na_filter=False)
            work_df.columns = work_df.columns.str.strip()
            debug['work_rows'] = len(work_df)

            # Memory optimization: keep only required columns early
            selected_work_cols = [
                "Customer", "Customer Name", "Number",
                "Date (UTC)", "Minimum Line Item Period Start (UTC)",
                "Maximum Line Item Period End (UTC)", "Inv Amount", "Inv Check"
            ]
            # if some selected cols missing, let later checks raise; to avoid KeyError here, filter intersection
            keep_cols = [c for c in selected_work_cols if c in work_df.columns]
            work_df = work_df[keep_cols].copy()

            # 3) open excel workbook bytes and list sheets
            status_text.text("📋 Analyzing Excel workbook...")
            progress_bar.progress(30)
            buf = revenue_file_buffer if hasattr(revenue_file_buffer, "read") else io.BytesIO(revenue_file_buffer)
            try:
                # For sheet detection use pandas with openpyxl engine
                buf.seek(0)
                xls = pd.ExcelFile(buf, engine='openpyxl')
                sheets = xls.sheet_names
                xls.close()
            except Exception:
                # fallback
                buf.seek(0)
                wb_tmp = load_workbook(buf, read_only=True)
                sheets = wb_tmp.sheetnames
                wb_tmp.close()

            # pick sheet
            if chosen_sheet:
                if chosen_sheet not in sheets:
                    raise ValueError(f"Chosen sheet '{chosen_sheet}' not in workbook.")
                rev_sheet = chosen_sheet
            else:
                if len(sheets) == 1:
                    rev_sheet = sheets[0]
                    st.info(f"📄 Using sheet: '{rev_sheet}'")
                else:
                    return None, {"sheets": sheets, "error": "multiple_sheets_choose_one"}

            # 4) find header row in revenue sheet preview
            status_text.text("🔍 Finding revenue sheet structure...")
            progress_bar.progress(40)
            rev_keys = [
                "Customer ID", "Account Name", "Inv. No.",
                "Date Issued", "Start Date", "End Date", "Collected"
            ]
            buf.seek(0)
            rev_hdr_idx = find_header_row_in_excel_bytes(buf, rev_sheet, rev_keys, preview_rows=preview_rows)

            # 5) load full revenue sheet as DataFrame
            status_text.text("📊 Loading revenue data...")
            progress_bar.progress(50)
            buf.seek(0)
            rev_df = pd.read_excel(buf, sheet_name=rev_sheet, header=rev_hdr_idx, dtype=str)
            rev_df.columns = rev_df.columns.str.strip()

            # 6) split table / remainder by first blank 'Inv. No.'
            status_text.text("🔢 Processing invoice data...")
            progress_bar.progress(60)
            inv_col = "Inv. No."
            if inv_col not in rev_df.columns:
                raise KeyError(f"'{inv_col}' not found in revenue sheet after header detection.")
            mask_blank = rev_df[inv_col].fillna("").astype(str).str.strip() == ""
            first_blank_idx = mask_blank[mask_blank].index.min() if mask_blank.any() else len(rev_df)
            table = rev_df.iloc[:first_blank_idx].copy()
            remainder = rev_df.iloc[first_blank_idx:].copy()
            debug['rev_table_rows'] = len(table)
            debug['rev_remainder_rows'] = len(remainder)
            debug['rev_hdr_index'] = rev_hdr_idx

            # 7) build mask on working df
            status_text.text("🎯 Filtering rows to append...")
            progress_bar.progress(70)
            inv_check_col = "Inv Check"
            inv_amount_col = "Inv Amount"
            if inv_check_col not in work_df.columns:
                raise KeyError(f"'{inv_check_col}' not found in working CSV columns: {list(work_df.columns)}")
            if inv_amount_col not in work_df.columns:
                raise KeyError(f"'{inv_amount_col}' not found in working CSV columns: {list(work_df.columns)}")

            work_df[inv_check_col] = work_df[inv_check_col].astype(str).fillna("").str.strip()
            inv_check_values_norm = {v.strip() for v in inv_check_values}
            mask_invcheck_na = work_df[inv_check_col].isin(inv_check_values_norm) | (work_df[inv_check_col] == "")
            work_df["_inv_amount_num"] = pd.to_numeric(work_df[inv_amount_col], errors="coerce").fillna(0.0)
            mask_invamount_nonzero = work_df["_inv_amount_num"] != 0.0
            mask_to_append = mask_invcheck_na & mask_invamount_nonzero
            debug['mask_invcheck_na'] = int(mask_invcheck_na.sum())
            debug['mask_invamount_nonzero'] = int(mask_invamount_nonzero.sum())
            debug['mask_both'] = int(mask_to_append.sum())

            # 8) select important columns and rename to revenue schema
            selected_work_cols_final = [
                "Customer", "Customer Name", "Number",
                "Date (UTC)", "Minimum Line Item Period Start (UTC)",
                "Maximum Line Item Period End (UTC)", "Inv Amount"
            ]
            for c in selected_work_cols_final:
                if c not in work_df.columns:
                    raise KeyError(f"Expected column '{c}' missing in working CSV")
            new_rows = work_df.loc[mask_to_append, selected_work_cols_final].copy()
            new_rows.columns = [
                "Customer ID", "Account Name", "Inv. No.",
                "Date Issued", "Start Date", "End Date", "Collected"
            ]

            # 9) de-dup
            status_text.text("🔄 Removing duplicates...")
            progress_bar.progress(80)
            existing = set(table["Inv. No."].astype(str).str.strip())
            before = len(new_rows)
            new_rows = new_rows[~new_rows["Inv. No."].astype(str).str.strip().isin(existing)].copy()
            after = len(new_rows)
            debug['new_candidates_before'] = before
            debug['new_candidates_after_dedup'] = after
            if after == 0:
                debug['message'] = "No new rows to insert after filters & dedupe."
                progress_bar.progress(100)
                status_text.text("✅ Complete - No new rows needed")
                return None, debug

            # 10) Insert rows into Excel workbook
            status_text.text(f"📝 Inserting {after} new rows into Excel...")
            progress_bar.progress(90)
            buf.seek(0)
            wb = load_workbook(buf, data_only=False, keep_vba=False)
            ws = wb[rev_sheet]
            header_excel_rownum = rev_hdr_idx + 1
            insert_at = header_excel_rownum + first_blank_idx + 1
            n_new = len(new_rows)
            ws.insert_rows(insert_at, amount=n_new)
            header_to_colidx = {name: idx+1 for idx, name in enumerate(rev_df.columns)}
            for row_offset, (_, r) in enumerate(new_rows.iterrows()):
                excel_row = insert_at + row_offset
                for col_name in rev_df.columns:
                    col_idx = header_to_colidx[col_name]
                    if col_name in new_rows.columns:
                        val = r[col_name]
                        if pd.isna(val):
                            val = ""
                        ws.cell(row=excel_row, column=col_idx, value=str(val))

            # Save workbook into bytes
            status_text.text("💾 Saving updated workbook...")
            progress_bar.progress(95)
            out_buf = io.BytesIO()
            wb.save(out_buf)
            wb.close()
            out_buf.seek(0)

            debug['insert_at_row'] = insert_at
            debug['inserted_rows'] = n_new
            debug['elapsed_s'] = time.time() - t0
            debug['message'] = f"Successfully inserted {n_new} rows at Excel row {insert_at} in sheet '{rev_sheet}'"

            progress_bar.progress(100)
            status_text.text("✅ Processing complete!")
            return out_buf.getvalue(), debug

        except Exception as e:
            progress_bar.progress(0)
            status_text.text("❌ Processing failed")
            raise e
        finally:
            # signal elapsed updater to stop and clean up UI placeholders after short delay
            _elapsed_stop.set()

            def _clear_progress():
                time.sleep(2)
                try:
                    progress_bar.empty()
                    status_text.empty()
                    elapsed_placeholder.empty()
                except Exception:
                    pass

            threading.Thread(target=_clear_progress, daemon=True).start()


    # -------------------
    # Streamlit UI (Tab content)
    # -------------------

    st.title("📊 Append Missing Working Rows -> Revenue Excel")

    st.markdown("""
    Upload your **Working CSV** and **Revenue Excel (.xlsx)** files. 
    
    ✨ **What this does:**
    - Finds rows where `Inv Check` is `#N/A` (or blank) and `Inv Amount` ≠ 0
    - Adds these rows to your Revenue Excel file
    - Removes duplicates automatically
    """)

    # Initialize session state
    if 'sheets_detected' not in st.session_state:
        st.session_state.sheets_detected = []
    if 'chosen_sheet' not in st.session_state:
        st.session_state.chosen_sheet = None

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("📄 Working CSV File")
        working_upload = st.file_uploader(
            "Upload Working CSV", 
            type=["csv", "txt"], 
            accept_multiple_files=False,
            help="CSV file containing working data with Inv Check and Inv Amount columns"
        )
        if working_upload is not None:
            file_size = len(working_upload.getvalue()) / (1024*1024)
            st.success(f"✅ File loaded: {working_upload.name} ({file_size:.1f} MB)")
            if file_size > 50:
                st.warning("⚠️ Large file detected. Processing may take longer.")
        use_local_working = st.checkbox("Or use server file path", value=False, key="local_work")
        working_local_path = ""
        if use_local_working:
            working_local_path = st.text_input("Local working CSV path", value="", key="work_path")

    with col2:
        st.subheader("📊 Revenue Excel File")
        revenue_upload = st.file_uploader(
            "Upload Revenue Excel (.xlsx)", 
            type=["xlsx", "xlsm", "xlsb"], 
            accept_multiple_files=False,
            help="Excel workbook containing revenue data"
        )
        if revenue_upload is not None:
            file_size = len(revenue_upload.getvalue()) / (1024*1024)
            st.success(f"✅ File loaded: {revenue_upload.name} ({file_size:.1f} MB)")
            if file_size > 50:
                st.warning("⚠️ Large file detected. Processing may take longer.")
        use_local_revenue = st.checkbox("Or use server file path", value=False, key="local_rev")
        revenue_local_path = ""
        if use_local_revenue:
            revenue_local_path = st.text_input("Local revenue Excel path", value="", key="rev_path")

    # Detect sheets when revenue file is uploaded or local path provided
    sheets_available = []
    if revenue_upload is not None or (use_local_revenue and revenue_local_path):
        try:
            with st.spinner("🔍 Detecting Excel sheets..."):
                if use_local_revenue and revenue_local_path:
                    xls = pd.ExcelFile(revenue_local_path, engine='openpyxl')
                    sheets_available = xls.sheet_names
                    xls.close()
                else:
                    revenue_bytes = revenue_upload.read()
                    revenue_upload.seek(0)
                    tmp_buf = io.BytesIO(revenue_bytes)
                    xls = pd.ExcelFile(tmp_buf, engine='openpyxl')
                    sheets_available = xls.sheet_names
                    xls.close()
                st.session_state.sheets_detected = sheets_available
        except Exception as e:
            st.error(f"❌ Error reading Excel file: {e}")
            st.session_state.sheets_detected = []

    # Show sheet selection
    if len(st.session_state.sheets_detected) > 1:
        st.subheader("📋 Sheet Selection")
        chosen_sheet = st.selectbox(
            "Choose Revenue sheet to update", 
            options=st.session_state.sheets_detected, 
            index=None,
            key="sheet_selector",
            help="Select the sheet containing your revenue data"
        )
        st.session_state.chosen_sheet = chosen_sheet
    elif len(st.session_state.sheets_detected) == 1:
        st.session_state.chosen_sheet = st.session_state.sheets_detected[0]
        st.success(f"📄 Auto-selected sheet: '{st.session_state.chosen_sheet}'")
    else:
        st.session_state.chosen_sheet = None

    # Determine if we can run
    can_run = False
    if use_local_working and use_local_revenue:
        can_run = bool(working_local_path and revenue_local_path)
    elif use_local_working:
        can_run = bool(working_local_path and revenue_upload is not None)
    elif use_local_revenue:
        can_run = bool(working_upload is not None and revenue_local_path)
    else:
        can_run = bool(working_upload is not None and revenue_upload is not None)

    if len(st.session_state.sheets_detected) > 1 and st.session_state.chosen_sheet is None:
        can_run = False
        st.warning("⚠️ Please select a sheet from the dropdown above.")

    st.markdown("---")

    if st.button("🚀 Process Files", disabled=not can_run, type="primary"):
        try:
            # Clear previous result
            if 'processing_result' in st.session_state:
                del st.session_state.processing_result

            # Prepare working & revenue sources
            if use_local_working:
                working_source = working_local_path
            else:
                working_source = io.BytesIO(working_upload.read())
                working_upload.seek(0)

            if use_local_revenue:
                revenue_source = revenue_local_path
            else:
                revenue_source = io.BytesIO(revenue_upload.read())
                revenue_upload.seek(0)

            st.info("🔄 Processing files... This may take a moment for large files.")

            # Call the function directly - it has its own timer built-in
            out_bytes, debug = append_missing_working_rows_streamlit(
                working_file_buffer=working_source,
                revenue_file_buffer=revenue_source,
                chosen_sheet=st.session_state.chosen_sheet,
            )

            # Store results
            st.session_state.processing_result = (out_bytes, debug)

        except Exception as e:
            st.error(f"❌ Processing failed: {str(e)}")
            if st.checkbox("Show detailed error", value=False):
                st.exception(e)

    # Display results if available
    if 'processing_result' in st.session_state:
        out_bytes, debug = st.session_state.processing_result

        if out_bytes is None:
            st.warning(f"ℹ️ {debug.get('message', 'No changes were made.')}")
        else:
            sheet_name = st.session_state.chosen_sheet or "Revenues"
            out_name = f"{sheet_name}_updated.xlsx"
            st.success(f"✅ {debug.get('message', 'Processing completed successfully!')}")
            col1, col2 = st.columns([2, 1])
            with col1:
                st.download_button(
                    "📥 Download Updated Excel File",
                    data=out_bytes,
                    file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            with col2:
                if st.button("🔄 Process Another File"):
                    for key in ['processing_result', 'sheets_detected', 'chosen_sheet']:
                        if key in st.session_state:
                            del st.session_state[key]
                    st.rerun()

        # Debug expander
        with st.expander("📊 Processing Details", expanded=False):
            st.json(debug)